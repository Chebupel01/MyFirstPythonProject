import sys
import io
import time
from datetime import datetime
from PyQt5 import uic
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidgetItem, QFileDialog
from PyQt5.QtGui import QPixmap, QTransform, QColor, QImage
from PyQt5.QtCore import QSize
from Resources_rc import *
from OtherFiles import *
from openpyxl import load_workbook
from openpyxl.writer.excel import save_workbook
import sqlite3
from string import Template
import math
import currency


ReLogWindow = """<?xml version="1.0" encoding="UTF-8"?>
<ui version="4.0">
 <class>Form</class>
 <widget class="QWidget" name="Form">
  <property name="geometry">
   <rect>
    <x>0</x>
    <y>0</y>
    <width>1274</width>
    <height>738</height>
   </rect>
  </property>
  <property name="minimumSize">
   <size>
    <width>600</width>
    <height>600</height>
   </size>
  </property>
  <property name="font">
   <font>
    <family>Arial</family>
    <weight>50</weight>
    <bold>false</bold>
   </font>
  </property>
  <property name="windowTitle">
   <string>Form</string>
  </property>
  <widget class="QWidget" name="widget" native="true">
   <property name="geometry">
    <rect>
     <x>450</x>
     <y>150</y>
     <width>500</width>
     <height>500</height>
    </rect>
   </property>
   <property name="minimumSize">
    <size>
     <width>500</width>
     <height>500</height>
    </size>
   </property>
   <property name="maximumSize">
    <size>
     <width>500</width>
     <height>500</height>
    </size>
   </property>
   <widget class="QLabel" name="label">
    <property name="geometry">
     <rect>
      <x>0</x>
      <y>0</y>
      <width>500</width>
      <height>500</height>
     </rect>
    </property>
    <property name="minimumSize">
     <size>
      <width>500</width>
      <height>500</height>
     </size>
    </property>
    <property name="maximumSize">
     <size>
      <width>500</width>
      <height>500</height>
     </size>
    </property>
    <property name="styleSheet">
     <string notr="true">border-image:url(:/Images/ReLogBackground.jpg);
border-radius: 20px;</string>
    </property>
    <property name="text">
     <string/>
    </property>
   </widget>
   <widget class="QPushButton" name="regButton">
    <property name="geometry">
     <rect>
      <x>40</x>
      <y>400</y>
      <width>181</width>
      <height>41</height>
     </rect>
    </property>
    <property name="font">
     <font>
      <family>Arial</family>
      <pointsize>10</pointsize>
      <weight>75</weight>
      <bold>true</bold>
     </font>
    </property>
    <property name="styleSheet">
     <string notr="true">QPushButton {
background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 10px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);
}

QPushButton:hover {
background-color: rgba(255,33,100,100);
}</string>
    </property>
    <property name="text">
     <string>Зарегистрироваться</string>
    </property>
   </widget>
   <widget class="QPushButton" name="logButton">
    <property name="geometry">
     <rect>
      <x>70</x>
      <y>350</y>
      <width>121</width>
      <height>41</height>
     </rect>
    </property>
    <property name="font">
     <font>
      <family>Arial</family>
      <pointsize>10</pointsize>
      <weight>75</weight>
      <italic>false</italic>
      <bold>true</bold>
     </font>
    </property>
    <property name="cursor">
     <cursorShape>PointingHandCursor</cursorShape>
    </property>
    <property name="styleSheet">
     <string notr="true">QPushButton {
background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 10px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);
}

QPushButton:hover {
background-color: rgba(255,33,100,100);
}</string>
    </property>
    <property name="text">
     <string>Войти</string>
    </property>
   </widget>
   <widget class="QLineEdit" name="enterPassword">
    <property name="geometry">
     <rect>
      <x>30</x>
      <y>265</y>
      <width>201</width>
      <height>45</height>
     </rect>
    </property>
    <property name="styleSheet">
     <string notr="true">QLineEdit {
background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 10px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);
}

QLineEdit:hover {
background-color: rgba(255,33,100,100);
}</string>
    </property>
   </widget>
   <widget class="QLineEdit" name="enterLogin">
    <property name="geometry">
     <rect>
      <x>30</x>
      <y>160</y>
      <width>201</width>
      <height>45</height>
     </rect>
    </property>
    <property name="styleSheet">
     <string notr="true">QLineEdit {
background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 10px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);
}

QLineEdit:hover {
background-color: rgba(255,33,100,100);
}</string>
    </property>
   </widget>
   <widget class="QLabel" name="label_2">
    <property name="geometry">
     <rect>
      <x>10</x>
      <y>110</y>
      <width>231</width>
      <height>371</height>
     </rect>
    </property>
    <property name="minimumSize">
     <size>
      <width>231</width>
      <height>371</height>
     </size>
    </property>
    <property name="maximumSize">
     <size>
      <width>231</width>
      <height>371</height>
     </size>
    </property>
    <property name="styleSheet">
     <string notr="true">background-color: rgba(0, 0, 0, 100);
border-radius: 10px;</string>
    </property>
    <property name="text">
     <string/>
    </property>
   </widget>
   <widget class="QLabel" name="label_3">
    <property name="geometry">
     <rect>
      <x>36</x>
      <y>220</y>
      <width>151</width>
      <height>41</height>
     </rect>
    </property>
    <property name="font">
     <font>
      <family>Arial</family>
      <pointsize>12</pointsize>
      <weight>75</weight>
      <bold>true</bold>
     </font>
    </property>
    <property name="styleSheet">
     <string notr="true">color:rgb(255, 255, 255)</string>
    </property>
    <property name="text">
     <string>Пароль</string>
    </property>
   </widget>
   <widget class="QLabel" name="label_4">
    <property name="geometry">
     <rect>
      <x>36</x>
      <y>110</y>
      <width>151</width>
      <height>41</height>
     </rect>
    </property>
    <property name="font">
     <font>
      <family>Arial</family>
      <pointsize>12</pointsize>
      <weight>75</weight>
      <bold>true</bold>
     </font>
    </property>
    <property name="styleSheet">
     <string notr="true">color:rgb(255, 255, 255)</string>
    </property>
    <property name="text">
     <string>Логин</string>
    </property>
   </widget>
   <widget class="QLabel" name="ReLogResult">
    <property name="geometry">
     <rect>
      <x>20</x>
      <y>440</y>
      <width>211</width>
      <height>31</height>
     </rect>
    </property>
    <property name="font">
     <font>
      <family>Arial</family>
      <pointsize>9</pointsize>
      <weight>75</weight>
      <bold>true</bold>
     </font>
    </property>
    <property name="styleSheet">
     <string notr="true">color:rgb(255, 255, 255)</string>
    </property>
    <property name="text">
     <string/>
    </property>
   </widget>
   <widget class="QPushButton" name="closeWindowButton">
    <property name="geometry">
     <rect>
      <x>440</x>
      <y>20</y>
      <width>41</width>
      <height>41</height>
     </rect>
    </property>
    <property name="font">
     <font>
      <pointsize>10</pointsize>
     </font>
    </property>
    <property name="styleSheet">
     <string notr="true">QPushButton {
background-color: rgba(255,33,100,150);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 10px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);
}

QPushButton:hover {
background-color: rgba(255,33,100,100);
}</string>
    </property>
    <property name="text">
     <string>X</string>
    </property>
   </widget>
   <widget class="QCheckBox" name="reminderButton">
    <property name="geometry">
     <rect>
      <x>40</x>
      <y>310</y>
      <width>181</width>
      <height>41</height>
     </rect>
    </property>
    <property name="font">
     <font>
      <family>Arial</family>
      <pointsize>10</pointsize>
      <weight>75</weight>
      <bold>true</bold>
     </font>
    </property>
    <property name="styleSheet">
     <string notr="true">QCheckBox {
background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 10px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);
}

QCheckBox:hover {
background-color: rgba(255,33,100,100);
}</string>
    </property>
    <property name="text">
     <string>Запомнить меня</string>
    </property>
   </widget>
   <zorder>label</zorder>
   <zorder>label_2</zorder>
   <zorder>regButton</zorder>
   <zorder>logButton</zorder>
   <zorder>enterPassword</zorder>
   <zorder>enterLogin</zorder>
   <zorder>label_3</zorder>
   <zorder>label_4</zorder>
   <zorder>ReLogResult</zorder>
   <zorder>closeWindowButton</zorder>
   <zorder>reminderButton</zorder>
  </widget>
 </widget>
 <resources>
  <include location="res.qrc"/>
  <include location="Resources.qrc"/>
 </resources>
 <connections/>
</ui>
"""

LoadWindow = """<?xml version="1.0" encoding="UTF-8"?>
<ui version="4.0">
 <class>Form</class>
 <widget class="QWidget" name="Form">
  <property name="geometry">
   <rect>
    <x>0</x>
    <y>0</y>
    <width>1003</width>
    <height>680</height>
   </rect>
  </property>
  <property name="font">
   <font>
    <family>Arial</family>
    <pointsize>12</pointsize>
    <weight>75</weight>
    <bold>true</bold>
   </font>
  </property>
  <property name="cursor">
   <cursorShape>PointingHandCursor</cursorShape>
  </property>
  <property name="windowTitle">
   <string>Form</string>
  </property>
  <widget class="QWidget" name="widget" native="true">
   <property name="geometry">
    <rect>
     <x>530</x>
     <y>280</y>
     <width>401</width>
     <height>281</height>
    </rect>
   </property>
   <widget class="QProgressBar" name="progressBar">
    <property name="geometry">
     <rect>
      <x>20</x>
      <y>50</y>
      <width>331</width>
      <height>41</height>
     </rect>
    </property>
    <property name="font">
     <font>
      <family>Arial</family>
      <pointsize>12</pointsize>
      <weight>75</weight>
      <bold>true</bold>
     </font>
    </property>
    <property name="focusPolicy">
     <enum>Qt::NoFocus</enum>
    </property>
    <property name="contextMenuPolicy">
     <enum>Qt::DefaultContextMenu</enum>
    </property>
    <property name="styleSheet">
     <string notr="true">QProgressBar{
background-color: rgb(124, 113, 116);
border-radius: 12px;
color: white;
text-align: center;
}

QProgressBar::chunk{
border-radius: 12px;
	background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:0, y2:0, stop:0 rgba(179, 65, 149), stop:1 rgba(179, 65, 244, 255));
}</string>
    </property>
    <property name="value">
     <number>0</number>
    </property>
   </widget>
   <widget class="QLabel" name="label">
    <property name="geometry">
     <rect>
      <x>6</x>
      <y>2</y>
      <width>361</width>
      <height>251</height>
     </rect>
    </property>
    <property name="styleSheet">
     <string notr="true">border-image:url(:/Images/LoadingBackground.jpg);
border-radius: 20px;</string>
    </property>
   </widget>
   <widget class="QLabel" name="label_2">
    <property name="geometry">
     <rect>
      <x>16</x>
      <y>103</y>
      <width>341</width>
      <height>121</height>
     </rect>
    </property>
    <property name="styleSheet">
     <string notr="true">background-color: rgba(0, 0, 0, 100);
border-radius: 10px;</string>
    </property>
    <property name="text">
     <string/>
    </property>
   </widget>
   <widget class="QPushButton" name="pushButton">
    <property name="geometry">
     <rect>
      <x>80</x>
      <y>170</y>
      <width>201</width>
      <height>41</height>
     </rect>
    </property>
    <property name="styleSheet">
     <string notr="true">QPushButton {
background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 10px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);
}

QPushButton:hover {
background-color: rgba(255,33,100,100);
}</string>
    </property>
    <property name="text">
     <string>Запустить</string>
    </property>
   </widget>
   <widget class="QCheckBox" name="checkBox">
    <property name="geometry">
     <rect>
      <x>50</x>
      <y>120</y>
      <width>261</width>
      <height>40</height>
     </rect>
    </property>
    <property name="font">
     <font>
      <family>Arial</family>
      <pointsize>10</pointsize>
      <weight>50</weight>
      <bold>false</bold>
     </font>
    </property>
    <property name="styleSheet">
     <string notr="true">QCheckBox {
background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 10px;
padding: 10px;
font-size: 10pt;
color: white;
}

QCheckBox:hover {
background-color: rgba(255,33,100,100);
}</string>
    </property>
    <property name="text">
     <string>Запустить по завершению загрузки</string>
    </property>
   </widget>
   <widget class="QLabel" name="label_3">
    <property name="geometry">
     <rect>
      <x>150</x>
      <y>20</y>
      <width>71</width>
      <height>21</height>
     </rect>
    </property>
    <property name="font">
     <font>
      <family>Arial</family>
      <pointsize>12</pointsize>
      <weight>75</weight>
      <bold>true</bold>
     </font>
    </property>
    <property name="styleSheet">
     <string notr="true">color: white;
background-color: rgba(0, 0, 0, 100);
border-radius: 10px;</string>
    </property>
    <property name="text">
     <string>Загрузка</string>
    </property>
   </widget>
   <zorder>label</zorder>
   <zorder>progressBar</zorder>
   <zorder>label_2</zorder>
   <zorder>pushButton</zorder>
   <zorder>checkBox</zorder>
   <zorder>label_3</zorder>
  </widget>
 </widget>
 <resources>
  <include location="res.qrc"/>
  <include location="Resources.qrc"/>
 </resources>
 <connections/>
</ui>
"""
MainWindowTemplate = """<?xml version="1.0" encoding="UTF-8"?>
<ui version="4.0">
 <class>Form</class>
 <widget class="QWidget" name="Form">
  <property name="geometry">
   <rect>
    <x>0</x>
    <y>0</y>
    <width>1030</width>
    <height>741</height>
   </rect>
  </property>
  <property name="minimumSize">
   <size>
    <width>1030</width>
    <height>741</height>
   </size>
  </property>
  <property name="maximumSize">
   <size>
    <width>16777215</width>
    <height>16777215</height>
   </size>
  </property>
  <property name="font">
   <font>
    <family>Arial</family>
    <pointsize>10</pointsize>
    <weight>75</weight>
    <bold>true</bold>
   </font>
  </property>
  <property name="windowTitle">
   <string>Form</string>
  </property>
  <property name="styleSheet">
   <string notr="true"/>
  </property>
  <widget class="QWidget" name="widget" native="true">
   <property name="geometry">
    <rect>
     <x>0</x>
     <y>0</y>
     <width>1031</width>
     <height>739</height>
    </rect>
   </property>
   <property name="minimumSize">
    <size>
     <width>1031</width>
     <height>0</height>
    </size>
   </property>
   <widget class="QLabel" name="background">
    <property name="geometry">
     <rect>
      <x>0</x>
      <y>0</y>
      <width>1031</width>
      <height>741</height>
     </rect>
    </property>
    <property name="styleSheet">
     <string notr="true">border-image:url(:/Images/BackgroudMainWindow.jpg);
border-radius: 20px;</string>
    </property>
    <property name="text">
     <string/>
    </property>
   </widget>
   <widget class="QPushButton" name="settings">
    <property name="geometry">
     <rect>
      <x>941</x>
      <y>35</y>
      <width>70</width>
      <height>70</height>
     </rect>
    </property>
    <property name="styleSheet">
     <string notr="true">QPushButton{
border-radius: 35px;
}

QPushButton:hover {
background-color: rgba(255,33,100,100);
}
</string>
    </property>
    <property name="text">
     <string/>
    </property>
   </widget>
   <widget class="QLabel" name="settingsLabel">
    <property name="geometry">
     <rect>
      <x>936</x>
      <y>30</y>
      <width>81</width>
      <height>81</height>
     </rect>
    </property>
    <property name="styleSheet">
     <string notr="true">border-image:url(:/Images/Без названия (2).png);
border-radius: 37px;</string>
    </property>
    <property name="text">
     <string/>
    </property>
   </widget>
   <widget class="QPushButton" name="currencyConverter">
    <property name="geometry">
     <rect>
      <x>739</x>
      <y>27</y>
      <width>181</width>
      <height>90</height>
     </rect>
    </property>
    <property name="minimumSize">
     <size>
      <width>90</width>
      <height>90</height>
     </size>
    </property>
    <property name="font">
     <font>
      <family>Arial</family>
      <pointsize>12</pointsize>
      <weight>75</weight>
      <bold>true</bold>
     </font>
    </property>
    <property name="styleSheet">
     <string notr="true">QPushButton{
border-radius: 45px;
background-color: rgb(44, 109, 168);
color: white;
}

QPushButton:hover {
background-color: rgba(255,33,100, 100);
}
</string>
    </property>
    <property name="text">
     <string>Конвертатор
валют</string>
    </property>
   </widget>
   <widget class="QLabel" name="currencyConverterLabel">
    <property name="geometry">
     <rect>
      <x>730</x>
      <y>20</y>
      <width>198</width>
      <height>104</height>
     </rect>
    </property>
    <property name="styleSheet">
     <string notr="true">background-color: rgb(255, 255, 255);
border-radius:50px</string>
    </property>
    <property name="text">
     <string/>
    </property>
   </widget>
   <widget class="QPushButton" name="expenseManager">
    <property name="geometry">
     <rect>
      <x>329</x>
      <y>27</y>
      <width>181</width>
      <height>90</height>
     </rect>
    </property>
    <property name="minimumSize">
     <size>
      <width>90</width>
      <height>90</height>
     </size>
    </property>
    <property name="font">
     <font>
      <family>Arial</family>
      <pointsize>12</pointsize>
      <weight>75</weight>
      <bold>true</bold>
     </font>
    </property>
    <property name="styleSheet">
     <string notr="true">QPushButton{
border-radius: 45px;
background-color: rgb(44, 109, 168);
color: white;
}

QPushButton:hover {
background-color: rgba(255,33,100,100);
}
</string>
    </property>
    <property name="text">
     <string>Менеджер
расходов</string>
    </property>
   </widget>
   <widget class="QLabel" name="expenseManagerLabel">
    <property name="geometry">
     <rect>
      <x>320</x>
      <y>20</y>
      <width>198</width>
      <height>104</height>
     </rect>
    </property>
    <property name="styleSheet">
     <string notr="true">background-color: rgb(255, 255, 255);
border-radius:50px</string>
    </property>
    <property name="text">
     <string/>
    </property>
   </widget>
   <widget class="QPushButton" name="revenueManager">
    <property name="geometry">
     <rect>
      <x>124</x>
      <y>27</y>
      <width>181</width>
      <height>90</height>
     </rect>
    </property>
    <property name="minimumSize">
     <size>
      <width>90</width>
      <height>90</height>
     </size>
    </property>
    <property name="font">
     <font>
      <family>Arial</family>
      <pointsize>12</pointsize>
      <weight>75</weight>
      <bold>true</bold>
     </font>
    </property>
    <property name="styleSheet">
     <string notr="true">QPushButton{
border-radius: 45px;
background-color: rgb(44, 109, 168);
color: white;
}

QPushButton:hover {
background-color: rgba(255,33,100,100);
}
</string>
    </property>
    <property name="text">
     <string>Менеджер
доходов</string>
    </property>
   </widget>
   <widget class="QLabel" name="revenueManagerLabel">
    <property name="geometry">
     <rect>
      <x>115</x>
      <y>20</y>
      <width>198</width>
      <height>104</height>
     </rect>
    </property>
    <property name="styleSheet">
     <string notr="true">background-color: rgb(255, 255, 255);
border-radius:50px</string>
    </property>
    <property name="text">
     <string/>
    </property>
   </widget>
   <widget class="QPushButton" name="goalPlanner">
    <property name="geometry">
     <rect>
      <x>534</x>
      <y>27</y>
      <width>181</width>
      <height>90</height>
     </rect>
    </property>
    <property name="minimumSize">
     <size>
      <width>90</width>
      <height>90</height>
     </size>
    </property>
    <property name="font">
     <font>
      <family>Arial</family>
      <pointsize>12</pointsize>
      <weight>75</weight>
      <bold>true</bold>
     </font>
    </property>
    <property name="styleSheet">
     <string notr="true">QPushButton{
border-radius: 45px;
background-color: rgb(44, 109, 168);
color: white;
}

QPushButton:hover {
background-color: rgba(255,33,100,100);
}
</string>
    </property>
    <property name="text">
     <string>Планировщик
целей</string>
    </property>
   </widget>
   <widget class="QLabel" name="goalPlannerManager">
    <property name="geometry">
     <rect>
      <x>525</x>
      <y>20</y>
      <width>198</width>
      <height>104</height>
     </rect>
    </property>
    <property name="styleSheet">
     <string notr="true">background-color: rgb(255, 255, 255);
border-radius:50px</string>
    </property>
    <property name="text">
     <string/>
    </property>
   </widget>
   <widget class="QPushButton" name="mainMenu">
    <property name="geometry">
     <rect>
      <x>16</x>
      <y>27</y>
      <width>85</width>
      <height>85</height>
     </rect>
    </property>
    <property name="autoFillBackground">
     <bool>false</bool>
    </property>
    <property name="styleSheet">
     <string notr="true">QPushButton{
border-radius: 42px;
}

QPushButton:hover {
background-color: rgba(255,33,100,100);
}
</string>
    </property>
    <property name="text">
     <string/>
    </property>
   </widget>
   <widget class="QLabel" name="mainMenuLabel">
    <property name="geometry">
     <rect>
      <x>10</x>
      <y>20</y>
      <width>100</width>
      <height>100</height>
     </rect>
    </property>
    <property name="styleSheet">
     <string notr="true">border-image:url(:/Images/Иконка приложения.png);
border-radius: 45px;</string>
    </property>
    <property name="text">
     <string/>
    </property>
   </widget>
   <widget class="QWidget" name="MainMenu" native="true">
    <property name="geometry">
     <rect>
      <x>30</x>
      <y>160</y>
      <width>961</width>
      <height>550</height>
     </rect>
    </property>
    <widget class="QLabel" name="MainMenuPlace">
     <property name="geometry">
      <rect>
       <x>0</x>
       <y>0</y>
       <width>961</width>
       <height>551</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">background-color: rgba(0, 0, 0, 150);
border-radius: 10px;</string>
     </property>
     <property name="text">
      <string/>
     </property>
    </widget>
    <widget class="QLabel" name="avatarFrame">
     <property name="geometry">
      <rect>
       <x>40</x>
       <y>50</y>
       <width>151</width>
       <height>151</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">border-image: url(:/Images/Рамка для аватарки.png)</string>
     </property>
     <property name="text">
      <string/>
     </property>
    </widget>
    <widget class="QLabel" name="avatar">
     <property name="geometry">
      <rect>
       <x>55</x>
       <y>65</y>
       <width>122</width>
       <height>122</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">border-image:url(:/Images/Иконка.png)</string>
     </property>
     <property name="text">
      <string/>
     </property>
    </widget>
    <widget class="QLabel" name="login">
     <property name="geometry">
      <rect>
       <x>20</x>
       <y>202</y>
       <width>190</width>
       <height>51</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>12</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">border-radius: 15px;
color: white;
background-color:rgba(44, 109, 168, 210);
border: 5px solid rgba(255, 255, 255, 250);</string>
     </property>
     <property name="text">
      <string/>
     </property>
    </widget>
    <widget class="QPushButton" name="editAvatar">
     <property name="geometry">
      <rect>
       <x>45</x>
       <y>272</y>
       <width>144</width>
       <height>51</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">QPushButton{
border-radius: 15px;
color: white;
background-color:rgba(44, 109, 168, 210);
border: 5px solid rgba(255, 255, 255, 250);
}

QPushButton:hover {
background-color: rgbargba(44, 109, 168, 150);
}</string>
     </property>
     <property name="text">
      <string>Изменить
аватарку</string>
     </property>
    </widget>
    <widget class="QLabel" name="balance">
     <property name="geometry">
      <rect>
       <x>450</x>
       <y>65</y>
       <width>491</width>
       <height>122</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>18</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">border-radius: 15px;
color: white;
background-color:rgba(44, 109, 168, 210);
border: 5px solid rgba(255, 255, 255, 250);</string>
     </property>
     <property name="text">
      <string/>
     </property>
    </widget>
    <widget class="QTableWidget" name="recentTransactions">
     <property name="geometry">
      <rect>
       <x>450</x>
       <y>260</y>
       <width>491</width>
       <height>271</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">color: black;
background-color:rgba(44, 109, 168, 210);
</string>
     </property>
    </widget>
    <widget class="QLabel" name="recentTransactionLabel">
     <property name="geometry">
      <rect>
       <x>450</x>
       <y>201</y>
       <width>491</width>
       <height>51</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>12</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="focusPolicy">
      <enum>Qt::NoFocus</enum>
     </property>
     <property name="styleSheet">
      <string notr="true">border-radius: 15px;
color: white;
background-color:rgba(44, 109, 168, 210);
border: 5px solid rgba(255, 255, 255, 250);</string>
     </property>
     <property name="text">
      <string>Последние 10 транзакций</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLabel" name="userInformation">
     <property name="geometry">
      <rect>
       <x>220</x>
       <y>65</y>
       <width>221</width>
       <height>465</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>12</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">border-radius: 15px;
color: white;
background-color:rgba(44, 109, 168, 210);
border: 5px solid rgba(255, 255, 255, 250);</string>
     </property>
     <property name="text">
      <string>&lt;html&gt;&lt;head/&gt;&lt;body&gt;&lt;p&gt;&lt;br/&gt;&lt;/p&gt;&lt;/body&gt;&lt;/html&gt;</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignLeading|Qt::AlignLeft|Qt::AlignTop</set>
     </property>
    </widget>
    <widget class="QLabel" name="MainMenuLabel">
     <property name="geometry">
      <rect>
       <x>320</x>
       <y>10</y>
       <width>331</width>
       <height>51</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>24</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Главное меню</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
   </widget>
   <widget class="QWidget" name="revenueManagerMenu" native="true">
    <property name="geometry">
     <rect>
      <x>30</x>
      <y>160</y>
      <width>961</width>
      <height>550</height>
     </rect>
    </property>
    <widget class="QLabel" name="revenueManagerPlace">
     <property name="geometry">
      <rect>
       <x>0</x>
       <y>0</y>
       <width>961</width>
       <height>551</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">background-color: rgba(0, 0, 0, 150);
border-radius: 10px;</string>
     </property>
     <property name="text">
      <string/>
     </property>
    </widget>
    <widget class="QLabel" name="revenueMenuLabel">
     <property name="geometry">
      <rect>
       <x>320</x>
       <y>10</y>
       <width>331</width>
       <height>51</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>24</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Менеджер доходов</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLabel" name="revenueHistoryLabel">
     <property name="geometry">
      <rect>
       <x>20</x>
       <y>70</y>
       <width>421</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>12</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="focusPolicy">
      <enum>Qt::NoFocus</enum>
     </property>
     <property name="styleSheet">
      <string notr="true">border-radius: 15px;
color: white;
background-color:rgba(34, 139, 34, 210);
border: 5px solid rgba(255, 255, 255, 250);</string>
     </property>
     <property name="text">
      <string>История доходов</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QTableWidget" name="revenueTransactions">
     <property name="geometry">
      <rect>
       <x>20</x>
       <y>150</y>
       <width>421</width>
       <height>381</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">color: black;
background-color:rgba(34, 139, 34, 210);
</string>
     </property>
    </widget>
    <widget class="QLabel" name="sortRevenueLabel">
     <property name="geometry">
      <rect>
       <x>36</x>
       <y>115</y>
       <width>111</width>
       <height>31</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Сортировать по</string>
     </property>
    </widget>
    <widget class="QComboBox" name="sortRevenueParameter">
     <property name="geometry">
      <rect>
       <x>150</x>
       <y>120</y>
       <width>121</width>
       <height>22</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">border: 5px solid rgba(255, 255, 255, 250);</string>
     </property>
     <item>
      <property name="text">
       <string>Дата</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Категория</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>От максимального</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>От минимального</string>
      </property>
     </item>
    </widget>
    <widget class="QPushButton" name="updateRevenueButton">
     <property name="geometry">
      <rect>
       <x>300</x>
       <y>117</y>
       <width>111</width>
       <height>26</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">QPushButton{
color: white;
background-color:rgba(34, 139, 34, 210);
border: 3px solid rgba(255, 255, 255, 250);
}

QPushButton:hover {
background-color:rgba(34, 139, 34, 100);
}</string>
     </property>
     <property name="text">
      <string>Обновить</string>
     </property>
    </widget>
    <widget class="QLabel" name="addRevenueLabel">
     <property name="geometry">
      <rect>
       <x>510</x>
       <y>70</y>
       <width>391</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>24</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Добавление доходов</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLabel" name="revenueSummaLabel">
     <property name="geometry">
      <rect>
       <x>600</x>
       <y>110</y>
       <width>211</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>12</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">border-radius: 15px;
color: white;
background-color:rgba(34, 139, 34, 210);
border: 5px solid rgba(255, 255, 255, 250);</string>
     </property>
     <property name="text">
      <string>Сумма</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLineEdit" name="revenueSummaEnter">
     <property name="geometry">
      <rect>
       <x>542</x>
       <y>159</y>
       <width>331</width>
       <height>41</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">color: rgba(255, 255, 255, 150);
background-color:rgba(34, 139, 34, 210);
border: 2px solid rgba(255, 255, 255, 250);</string>
     </property>
     <property name="text">
      <string/>
     </property>
     <property name="placeholderText">
      <string>Например (3125)</string>
     </property>
    </widget>
    <widget class="QLineEdit" name="revenueCategoryEnter">
     <property name="geometry">
      <rect>
       <x>542</x>
       <y>259</y>
       <width>331</width>
       <height>41</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">color: rgba(255, 255, 255, 150);
background-color:rgba(34, 139, 34, 210);
border: 2px solid rgba(255, 255, 255, 250);</string>
     </property>
     <property name="text">
      <string/>
     </property>
     <property name="placeholderText">
      <string>Например (Зарплата, перевод)</string>
     </property>
    </widget>
    <widget class="QLabel" name="revenueCategoryLabel">
     <property name="geometry">
      <rect>
       <x>600</x>
       <y>210</y>
       <width>211</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>12</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">border-radius: 15px;
color: white;
background-color:rgba(34, 139, 34, 210);
border: 5px solid rgba(255, 255, 255, 250);</string>
     </property>
     <property name="text">
      <string>Категория</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLineEdit" name="revenueSourceEnter">
     <property name="geometry">
      <rect>
       <x>542</x>
       <y>359</y>
       <width>331</width>
       <height>41</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">color: rgba(255, 255, 255, 150);
background-color:rgba(34, 139, 34, 210);
border: 2px solid rgba(255, 255, 255, 250);</string>
     </property>
     <property name="text">
      <string/>
     </property>
     <property name="placeholderText">
      <string>Например (Работа, бизнес)</string>
     </property>
    </widget>
    <widget class="QLabel" name="revenueSourceLabel">
     <property name="geometry">
      <rect>
       <x>600</x>
       <y>310</y>
       <width>211</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>12</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">border-radius: 15px;
color: white;
background-color:rgba(34, 139, 34, 210);
border: 5px solid rgba(255, 255, 255, 250);</string>
     </property>
     <property name="text">
      <string>Источник дохода</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QPushButton" name="addRevenue">
     <property name="geometry">
      <rect>
       <x>610</x>
       <y>410</y>
       <width>191</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>14</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">QPushButton{
color: white;
background-color:rgba(34, 139, 34, 210);
border: 3px solid rgba(255, 255, 255, 250);
}

QPushButton:hover {
background-color:rgba(34, 139, 34, 100);
}</string>
     </property>
     <property name="text">
      <string>Добавить</string>
     </property>
    </widget>
    <widget class="QLabel" name="addRevenueErrorLabel">
     <property name="geometry">
      <rect>
       <x>670</x>
       <y>20</y>
       <width>271</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>TextLabel</string>
     </property>
    </widget>
   </widget>
   <widget class="QWidget" name="expenseManagerMenu" native="true">
    <property name="geometry">
     <rect>
      <x>30</x>
      <y>160</y>
      <width>961</width>
      <height>550</height>
     </rect>
    </property>
    <widget class="QLabel" name="expenseManagerPlace">
     <property name="geometry">
      <rect>
       <x>0</x>
       <y>0</y>
       <width>961</width>
       <height>551</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">background-color: rgba(0, 0, 0, 150);
border-radius: 10px;</string>
     </property>
     <property name="text">
      <string/>
     </property>
    </widget>
    <widget class="QLabel" name="expenseMenuLabel">
     <property name="geometry">
      <rect>
       <x>320</x>
       <y>10</y>
       <width>331</width>
       <height>50</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>24</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Менеджер расходов</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLabel" name="expenseHistoryLabel">
     <property name="geometry">
      <rect>
       <x>20</x>
       <y>70</y>
       <width>421</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>12</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="focusPolicy">
      <enum>Qt::NoFocus</enum>
     </property>
     <property name="styleSheet">
      <string notr="true">border-radius: 15px;
color: white;
background-color:rgba(178, 34, 34, 210);
border: 5px solid rgba(255, 255, 255, 250);</string>
     </property>
     <property name="text">
      <string>История расходов</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QTableWidget" name="expenseTransactions">
     <property name="geometry">
      <rect>
       <x>20</x>
       <y>150</y>
       <width>421</width>
       <height>381</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">color: black;
background-color:rgba(178, 34, 34, 210);
</string>
     </property>
    </widget>
    <widget class="QLabel" name="sortExpenseLabel">
     <property name="geometry">
      <rect>
       <x>36</x>
       <y>115</y>
       <width>111</width>
       <height>31</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Сортировать по</string>
     </property>
    </widget>
    <widget class="QComboBox" name="sortExpenseParameter">
     <property name="geometry">
      <rect>
       <x>150</x>
       <y>120</y>
       <width>141</width>
       <height>22</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">border: 5px solid rgba(255, 255, 255, 250);</string>
     </property>
     <property name="frame">
      <bool>true</bool>
     </property>
     <item>
      <property name="text">
       <string>Дата</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Категория</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>От максимальной суммы</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>От минимальной суммы</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Источник </string>
      </property>
     </item>
    </widget>
    <widget class="QPushButton" name="updateExpenseButton">
     <property name="geometry">
      <rect>
       <x>300</x>
       <y>118</y>
       <width>111</width>
       <height>26</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">QPushButton{
color: white;
background-color:rgba(178, 34, 34, 210);
border: 3px solid rgba(255, 255, 255, 250);
}

QPushButton:hover {
background-color:rgba(178, 34, 34, 100);
}</string>
     </property>
     <property name="text">
      <string>Обновить</string>
     </property>
    </widget>
    <widget class="QLabel" name="addExpenseLabel">
     <property name="geometry">
      <rect>
       <x>510</x>
       <y>70</y>
       <width>391</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>24</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Добавление расходов</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLabel" name="expenseSummaLabel">
     <property name="geometry">
      <rect>
       <x>600</x>
       <y>110</y>
       <width>211</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>12</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">border-radius: 15px;
color: white;
background-color:rgba(178, 34, 34, 210);
border: 5px solid rgba(255, 255, 255, 250);</string>
     </property>
     <property name="text">
      <string>Сумма</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLineEdit" name="expenseSummaEnter">
     <property name="geometry">
      <rect>
       <x>542</x>
       <y>159</y>
       <width>331</width>
       <height>41</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">color: rgba(255, 255, 255, 150);
background-color:rgba(178, 34, 34, 210);
border: 2px solid rgba(255, 255, 255, 250);</string>
     </property>
     <property name="text">
      <string/>
     </property>
     <property name="placeholderText">
      <string>Например (3125)</string>
     </property>
    </widget>
    <widget class="QLineEdit" name="expenseCategoryEnter">
     <property name="geometry">
      <rect>
       <x>542</x>
       <y>259</y>
       <width>331</width>
       <height>41</height>
      </rect>
     </property>
     <property name="accessibleName">
      <string/>
     </property>
     <property name="autoFillBackground">
      <bool>false</bool>
     </property>
     <property name="styleSheet">
      <string notr="true">color: rgba(255, 255, 255, 150);
background-color:rgba(178, 34, 34, 210);
border: 2px solid rgba(255, 255, 255, 250);</string>
     </property>
     <property name="inputMask">
      <string/>
     </property>
     <property name="text">
      <string notr="true"/>
     </property>
     <property name="placeholderText">
      <string>Например (Покупка, перевод)</string>
     </property>
    </widget>
    <widget class="QLabel" name="expenseCategoryLabel">
     <property name="geometry">
      <rect>
       <x>600</x>
       <y>210</y>
       <width>211</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>12</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">border-radius: 15px;
color: white;
background-color:rgba(178, 34, 34, 210);
border: 5px solid rgba(255, 255, 255, 250);</string>
     </property>
     <property name="text">
      <string>Категория</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLineEdit" name="expenseSourceEnter">
     <property name="geometry">
      <rect>
       <x>542</x>
       <y>359</y>
       <width>331</width>
       <height>41</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">color: rgba(255, 255, 255, 150);
background-color:rgba(178, 34, 34, 210);
border: 2px solid rgba(255, 255, 255, 250);</string>
     </property>
     <property name="text">
      <string/>
     </property>
     <property name="placeholderText">
      <string>Например (Интернет-магазин, магазин одежды)</string>
     </property>
    </widget>
    <widget class="QLabel" name="expenseSourceLabel">
     <property name="geometry">
      <rect>
       <x>600</x>
       <y>310</y>
       <width>211</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>12</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">border-radius: 15px;
color: white;
background-color:rgba(178, 34, 34, 210);
border: 5px solid rgba(255, 255, 255, 250);</string>
     </property>
     <property name="text">
      <string>Источник расходов</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QPushButton" name="addExpense">
     <property name="geometry">
      <rect>
       <x>610</x>
       <y>410</y>
       <width>191</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>14</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">QPushButton{
color: white;
background-color:rgba(178, 34, 34, 210);
border: 3px solid rgba(255, 255, 255, 250);
}

QPushButton:hover {
background-color:rgba(178, 34, 34, 100);
}</string>
     </property>
     <property name="text">
      <string>Добавить</string>
     </property>
    </widget>
    <widget class="QLabel" name="addExpenseErrorLabel">
     <property name="geometry">
      <rect>
       <x>670</x>
       <y>20</y>
       <width>271</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>TextLabel</string>
     </property>
    </widget>
   </widget>
   <widget class="QWidget" name="goalPlannerMenu" native="true">
    <property name="geometry">
     <rect>
      <x>30</x>
      <y>160</y>
      <width>961</width>
      <height>551</height>
     </rect>
    </property>
    <widget class="QLabel" name="goalPlannerPlace">
     <property name="geometry">
      <rect>
       <x>0</x>
       <y>0</y>
       <width>961</width>
       <height>551</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">background-color: rgba(0, 0, 0, 150);
border-radius: 10px;</string>
     </property>
     <property name="text">
      <string/>
     </property>
    </widget>
    <widget class="QLabel" name="goalPlannerLabel">
     <property name="geometry">
      <rect>
       <x>320</x>
       <y>10</y>
       <width>330</width>
       <height>50</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>24</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Планировщик целей</string>
     </property>
    </widget>
    <widget class="QProgressBar" name="goalProgress1">
     <property name="geometry">
      <rect>
       <x>70</x>
       <y>100</y>
       <width>311</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">QProgressBar{
background-color: rgb(124, 113, 116);
border-radius: 12px;
color: white;
text-align: center;
border: 2px solid rgba(255, 255, 255, 250);
}

QProgressBar::chunk{
border-radius: 12px;
	background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:0, y2:0, stop:0 rgba(179, 65, 149), stop:1 rgba(179, 65, 244, 255));
}</string>
     </property>
     <property name="value">
      <number>0</number>
     </property>
    </widget>
    <widget class="QLabel" name="goalProgressLabel1">
     <property name="geometry">
      <rect>
       <x>70</x>
       <y>60</y>
       <width>311</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Не используется</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLabel" name="goalProgressLabel2">
     <property name="geometry">
      <rect>
       <x>70</x>
       <y>170</y>
       <width>311</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Не используется</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QProgressBar" name="goalProgress2">
     <property name="geometry">
      <rect>
       <x>70</x>
       <y>210</y>
       <width>311</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">QProgressBar{
background-color: rgb(124, 113, 116);
border-radius: 12px;
color: white;
text-align: center;
border: 2px solid rgba(255, 255, 255, 250);
}

QProgressBar::chunk{
border-radius: 12px;
	background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:0, y2:0, stop:0 rgba(179, 65, 149), stop:1 rgba(179, 65, 244, 255));
}</string>
     </property>
     <property name="value">
      <number>0</number>
     </property>
    </widget>
    <widget class="QLabel" name="goalProgressLabel3">
     <property name="geometry">
      <rect>
       <x>70</x>
       <y>280</y>
       <width>311</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Не используется</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QProgressBar" name="goalProgress3">
     <property name="geometry">
      <rect>
       <x>70</x>
       <y>320</y>
       <width>311</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">QProgressBar{
background-color: rgb(124, 113, 116);
border-radius: 12px;
color: white;
text-align: center;
border: 2px solid rgba(255, 255, 255, 250);
}

QProgressBar::chunk{
border-radius: 12px;
	background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:0, y2:0, stop:0 rgba(179, 65, 149), stop:1 rgba(179, 65, 244, 255));
}</string>
     </property>
     <property name="value">
      <number>0</number>
     </property>
    </widget>
    <widget class="QLabel" name="goalProgressLabel4">
     <property name="geometry">
      <rect>
       <x>70</x>
       <y>390</y>
       <width>311</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Не используется</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QProgressBar" name="goalProgress4">
     <property name="geometry">
      <rect>
       <x>70</x>
       <y>430</y>
       <width>311</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">QProgressBar{
background-color: rgb(124, 113, 116);
border-radius: 12px;
color: white;
text-align: center;
border: 2px solid rgba(255, 255, 255, 250);
}

QProgressBar::chunk{
border-radius: 12px;
	background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:0, y2:0, stop:0 rgba(179, 65, 149), stop:1 rgba(179, 65, 244, 255));
}</string>
     </property>
     <property name="value">
      <number>0</number>
     </property>
    </widget>
    <widget class="QLabel" name="goalProgressLabel5">
     <property name="geometry">
      <rect>
       <x>580</x>
       <y>60</y>
       <width>311</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Не используется</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QProgressBar" name="goalProgress5">
     <property name="geometry">
      <rect>
       <x>580</x>
       <y>100</y>
       <width>311</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">QProgressBar{
background-color: rgb(124, 113, 116);
border-radius: 12px;
color: white;
text-align: center;
border: 2px solid rgba(255, 255, 255, 250);
}

QProgressBar::chunk{
border-radius: 12px;
	background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:0, y2:0, stop:0 rgba(179, 65, 149), stop:1 rgba(179, 65, 244, 255));
}</string>
     </property>
     <property name="value">
      <number>0</number>
     </property>
    </widget>
    <widget class="QLabel" name="goalProgressLabel6">
     <property name="geometry">
      <rect>
       <x>580</x>
       <y>170</y>
       <width>311</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Не используется</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QProgressBar" name="goalProgress6">
     <property name="geometry">
      <rect>
       <x>580</x>
       <y>210</y>
       <width>311</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">QProgressBar{
background-color: rgb(124, 113, 116);
border-radius: 12px;
color: white;
text-align: center;
border: 2px solid rgba(255, 255, 255, 250);
}

QProgressBar::chunk{
border-radius: 12px;
	background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:0, y2:0, stop:0 rgba(179, 65, 149), stop:1 rgba(179, 65, 244, 255));
}</string>
     </property>
     <property name="value">
      <number>0</number>
     </property>
    </widget>
    <widget class="QLabel" name="goalProgressLabel7">
     <property name="geometry">
      <rect>
       <x>580</x>
       <y>280</y>
       <width>311</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Не используется</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QProgressBar" name="goalProgress7">
     <property name="geometry">
      <rect>
       <x>580</x>
       <y>320</y>
       <width>311</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">QProgressBar{
background-color: rgb(124, 113, 116);
border-radius: 12px;
color: white;
text-align: center;
border: 2px solid rgba(255, 255, 255, 250);
}

QProgressBar::chunk{
border-radius: 12px;
	background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:0, y2:0, stop:0 rgba(179, 65, 149), stop:1 rgba(179, 65, 244, 255));
}</string>
     </property>
     <property name="value">
      <number>0</number>
     </property>
    </widget>
    <widget class="QLabel" name="goalProgressLabel8">
     <property name="geometry">
      <rect>
       <x>580</x>
       <y>390</y>
       <width>311</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Не используется</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QProgressBar" name="goalProgress8">
     <property name="geometry">
      <rect>
       <x>580</x>
       <y>430</y>
       <width>311</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">QProgressBar{
background-color: rgb(124, 113, 116);
border-radius: 12px;
color: white;
text-align: center;
border: 2px solid rgba(255, 255, 255, 250);
}

QProgressBar::chunk{
border-radius: 12px;
	background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:0, y2:0, stop:0 rgba(179, 65, 149), stop:1 rgba(179, 65, 244, 255));
}</string>
     </property>
     <property name="value">
      <number>0</number>
     </property>
    </widget>
    <widget class="QPushButton" name="updateGoals">
     <property name="geometry">
      <rect>
       <x>410</x>
       <y>102</y>
       <width>141</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">QPushButton{
border-radius: 10px;
background-color:qlineargradient(spread:reflect, x1:0, y1:0.011, x2:1, y2:0, stop:0 rgba(189, 36, 126, 243), stop:1 rgba(255, 255, 255, 255));
color: white;
border: 2px solid rgba(255, 33, 100, 230);
}

QPushButton:hover {
background-color: qlineargradient(spread:reflect, x1:0, y1:0.011, x2:1, y2:0, stop:0 rgba(189, 36, 126, 200), stop:1 rgba(255, 255, 255, 200));
}
</string>
     </property>
     <property name="text">
      <string>Обновить</string>
     </property>
    </widget>
    <widget class="QPushButton" name="createGoal">
     <property name="geometry">
      <rect>
       <x>410</x>
       <y>180</y>
       <width>141</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">QPushButton{
border-radius: 10px;
background-color:qlineargradient(spread:reflect, x1:0, y1:0.011, x2:1, y2:0, stop:0 rgba(189, 36, 126, 243), stop:1 rgba(255, 255, 255, 255));
color: white;
border: 2px solid rgba(255, 33, 100, 230);
}

QPushButton:hover {
background-color: qlineargradient(spread:reflect, x1:0, y1:0.011, x2:1, y2:0, stop:0 rgba(189, 36, 126, 200), stop:1 rgba(255, 255, 255, 200));
}

</string>
     </property>
     <property name="text">
      <string>Создать</string>
     </property>
    </widget>
    <widget class="QPushButton" name="editGoal">
     <property name="geometry">
      <rect>
       <x>410</x>
       <y>260</y>
       <width>141</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">QPushButton{
border-radius: 10px;
background-color:qlineargradient(spread:reflect, x1:0, y1:0.011, x2:1, y2:0, stop:0 rgba(189, 36, 126, 243), stop:1 rgba(255, 255, 255, 255));
color: white;
border: 2px solid rgba(255, 33, 100, 230);
}

QPushButton:hover {
background-color: qlineargradient(spread:reflect, x1:0, y1:0.011, x2:1, y2:0, stop:0 rgba(189, 36, 126, 200), stop:1 rgba(255, 255, 255, 200));
}

</string>
     </property>
     <property name="text">
      <string>Редактировать</string>
     </property>
    </widget>
    <widget class="QWidget" name="addGoal" native="true">
     <property name="geometry">
      <rect>
       <x>329</x>
       <y>49</y>
       <width>321</width>
       <height>461</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">border-radius: 10px;
background-color: qlineargradient(spread:reflect, x1:0, y1:0.011, x2:1, y2:0, stop:0 rgba(189, 36, 126, 243), stop:1 rgba(255, 255, 255, 255));</string>
     </property>
     <widget class="QLabel" name="addGoalLabel">
      <property name="geometry">
       <rect>
        <x>0</x>
        <y>30</y>
        <width>321</width>
        <height>41</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <family>Arial Black</family>
        <pointsize>14</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">color: white;</string>
      </property>
      <property name="text">
       <string>Создать цель</string>
      </property>
      <property name="alignment">
       <set>Qt::AlignCenter</set>
      </property>
     </widget>
     <widget class="QLabel" name="goalNameLabel">
      <property name="geometry">
       <rect>
        <x>0</x>
        <y>70</y>
        <width>321</width>
        <height>41</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <family>Arial Black</family>
        <pointsize>14</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">color: white;</string>
      </property>
      <property name="text">
       <string>Название</string>
      </property>
      <property name="alignment">
       <set>Qt::AlignLeading|Qt::AlignLeft|Qt::AlignVCenter</set>
      </property>
     </widget>
     <widget class="QLabel" name="goalTargetLabel">
      <property name="geometry">
       <rect>
        <x>0</x>
        <y>150</y>
        <width>321</width>
        <height>41</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <family>Arial Black</family>
        <pointsize>14</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">color: white;</string>
      </property>
      <property name="text">
       <string>Цель</string>
      </property>
      <property name="alignment">
       <set>Qt::AlignLeading|Qt::AlignLeft|Qt::AlignVCenter</set>
      </property>
     </widget>
     <widget class="QLabel" name="selectColorLabel">
      <property name="geometry">
       <rect>
        <x>0</x>
        <y>240</y>
        <width>321</width>
        <height>41</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <family>Arial Black</family>
        <pointsize>14</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">color: white;</string>
      </property>
      <property name="text">
       <string>Цвет</string>
      </property>
      <property name="alignment">
       <set>Qt::AlignLeading|Qt::AlignLeft|Qt::AlignVCenter</set>
      </property>
     </widget>
     <widget class="QLineEdit" name="enterGoalName">
      <property name="geometry">
       <rect>
        <x>12</x>
        <y>115</y>
        <width>301</width>
        <height>41</height>
       </rect>
      </property>
      <property name="styleSheet">
       <string notr="true">QLineEdit {
background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 3px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);
}

</string>
      </property>
     </widget>
     <widget class="QLineEdit" name="enterGoalTarget">
      <property name="geometry">
       <rect>
        <x>10</x>
        <y>195</y>
        <width>301</width>
        <height>41</height>
       </rect>
      </property>
      <property name="styleSheet">
       <string notr="true">QLineEdit {
background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 3px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);
}

</string>
      </property>
     </widget>
     <widget class="QComboBox" name="selectColor">
      <property name="geometry">
       <rect>
        <x>20</x>
        <y>281</y>
        <width>171</width>
        <height>41</height>
       </rect>
      </property>
      <property name="styleSheet">
       <string notr="true">background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 3px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);</string>
      </property>
      <item>
       <property name="text">
        <string>Синий</string>
       </property>
      </item>
      <item>
       <property name="text">
        <string>Красный</string>
       </property>
      </item>
      <item>
       <property name="text">
        <string>Жёлтый</string>
       </property>
      </item>
      <item>
       <property name="text">
        <string>Чёрный</string>
       </property>
      </item>
      <item>
       <property name="text">
        <string>Фиолетовый</string>
       </property>
      </item>
      <item>
       <property name="text">
        <string>Зелёный</string>
       </property>
      </item>
      <item>
       <property name="text">
        <string>Оранжевый</string>
       </property>
      </item>
      <item>
       <property name="text">
        <string>Розовый</string>
       </property>
      </item>
     </widget>
     <widget class="QPushButton" name="addGoalButton">
      <property name="geometry">
       <rect>
        <x>60</x>
        <y>330</y>
        <width>211</width>
        <height>51</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <family>Arial</family>
        <pointsize>10</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="cursor">
       <cursorShape>PointingHandCursor</cursorShape>
      </property>
      <property name="styleSheet">
       <string notr="true">QPushButton {
background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 10px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);
}

QPushButton:hover {
background-color: rgba(255,33,100,100);
}</string>
      </property>
      <property name="text">
       <string>Создать</string>
      </property>
     </widget>
     <widget class="QPushButton" name="closeAddGoalWindowButton">
      <property name="geometry">
       <rect>
        <x>270</x>
        <y>10</y>
        <width>41</width>
        <height>41</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <pointsize>10</pointsize>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">QPushButton {
background-color: rgba(255,33,100,150);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 10px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);
}

QPushButton:hover {
background-color: rgba(255,33,100,100);
}</string>
      </property>
      <property name="text">
       <string>X</string>
      </property>
     </widget>
     <widget class="QLabel" name="addGoalResult">
      <property name="geometry">
       <rect>
        <x>0</x>
        <y>390</y>
        <width>321</width>
        <height>41</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <family>Arial</family>
        <pointsize>12</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">color:white;</string>
      </property>
      <property name="text">
       <string/>
      </property>
     </widget>
    </widget>
    <widget class="QLabel" name="goalNumber_1">
     <property name="geometry">
      <rect>
       <x>20</x>
       <y>100</y>
       <width>41</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>24</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color:white;</string>
     </property>
     <property name="text">
      <string>1</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLabel" name="goalNumber_2">
     <property name="geometry">
      <rect>
       <x>20</x>
       <y>210</y>
       <width>41</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>24</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color:white;</string>
     </property>
     <property name="text">
      <string>2</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLabel" name="goalNumber_3">
     <property name="geometry">
      <rect>
       <x>20</x>
       <y>320</y>
       <width>41</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>24</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color:white;</string>
     </property>
     <property name="text">
      <string>3</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLabel" name="goalNumber_4">
     <property name="geometry">
      <rect>
       <x>20</x>
       <y>430</y>
       <width>41</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>24</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color:white;</string>
     </property>
     <property name="text">
      <string>4</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLabel" name="goalNumber_5">
     <property name="geometry">
      <rect>
       <x>900</x>
       <y>100</y>
       <width>41</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>24</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color:white;</string>
     </property>
     <property name="text">
      <string>5</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLabel" name="goalNumber_6">
     <property name="geometry">
      <rect>
       <x>900</x>
       <y>210</y>
       <width>41</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>24</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color:white;</string>
     </property>
     <property name="text">
      <string>6</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLabel" name="goalNumber_7">
     <property name="geometry">
      <rect>
       <x>900</x>
       <y>320</y>
       <width>41</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>24</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color:white;</string>
     </property>
     <property name="text">
      <string>7</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLabel" name="goalNumber_8">
     <property name="geometry">
      <rect>
       <x>900</x>
       <y>430</y>
       <width>41</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>24</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color:white;</string>
     </property>
     <property name="text">
      <string>8</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QPushButton" name="addProgress">
     <property name="geometry">
      <rect>
       <x>410</x>
       <y>340</y>
       <width>141</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>10</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">QPushButton{
border-radius: 10px;
background-color:qlineargradient(spread:reflect, x1:0, y1:0.011, x2:1, y2:0, stop:0 rgba(189, 36, 126, 243), stop:1 rgba(255, 255, 255, 255));
color: white;
border: 2px solid rgba(255, 33, 100, 230);
}

QPushButton:hover {
background-color: qlineargradient(spread:reflect, x1:0, y1:0.011, x2:1, y2:0, stop:0 rgba(189, 36, 126, 200), stop:1 rgba(255, 255, 255, 200));
}

</string>
     </property>
     <property name="text">
      <string>Добавить
прогреcc  цели</string>
     </property>
    </widget>
    <widget class="QWidget" name="editGoalMenu" native="true">
     <property name="geometry">
      <rect>
       <x>329</x>
       <y>49</y>
       <width>321</width>
       <height>461</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">border-radius: 10px;
background-color: qlineargradient(spread:reflect, x1:0, y1:0.011, x2:1, y2:0, stop:0 rgba(189, 36, 126, 243), stop:1 rgba(255, 255, 255, 255));</string>
     </property>
     <widget class="QLabel" name="editGoalLabel">
      <property name="geometry">
       <rect>
        <x>0</x>
        <y>30</y>
        <width>321</width>
        <height>41</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <family>Arial Black</family>
        <pointsize>14</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">color: white;</string>
      </property>
      <property name="text">
       <string>Редактировать цель</string>
      </property>
      <property name="alignment">
       <set>Qt::AlignCenter</set>
      </property>
     </widget>
     <widget class="QLabel" name="goalEditNameLabel">
      <property name="geometry">
       <rect>
        <x>0</x>
        <y>100</y>
        <width>321</width>
        <height>41</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <family>Arial Black</family>
        <pointsize>14</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">color: white;</string>
      </property>
      <property name="text">
       <string>Название</string>
      </property>
      <property name="alignment">
       <set>Qt::AlignLeading|Qt::AlignLeft|Qt::AlignVCenter</set>
      </property>
     </widget>
     <widget class="QLabel" name="goalEditTargetLabel">
      <property name="geometry">
       <rect>
        <x>0</x>
        <y>180</y>
        <width>321</width>
        <height>41</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <family>Arial Black</family>
        <pointsize>14</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">color: white;</string>
      </property>
      <property name="text">
       <string>Цель</string>
      </property>
      <property name="alignment">
       <set>Qt::AlignLeading|Qt::AlignLeft|Qt::AlignVCenter</set>
      </property>
     </widget>
     <widget class="QLabel" name="selectEditColorLabel">
      <property name="geometry">
       <rect>
        <x>0</x>
        <y>265</y>
        <width>321</width>
        <height>41</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <family>Arial Black</family>
        <pointsize>14</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">color: white;</string>
      </property>
      <property name="text">
       <string>Цвет</string>
      </property>
      <property name="alignment">
       <set>Qt::AlignLeading|Qt::AlignLeft|Qt::AlignVCenter</set>
      </property>
     </widget>
     <widget class="QLineEdit" name="enterEditGoalName">
      <property name="geometry">
       <rect>
        <x>12</x>
        <y>145</y>
        <width>301</width>
        <height>41</height>
       </rect>
      </property>
      <property name="styleSheet">
       <string notr="true">QLineEdit {
background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 3px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);
}

</string>
      </property>
     </widget>
     <widget class="QLineEdit" name="enterEditGoalTarget">
      <property name="geometry">
       <rect>
        <x>10</x>
        <y>230</y>
        <width>301</width>
        <height>41</height>
       </rect>
      </property>
      <property name="styleSheet">
       <string notr="true">QLineEdit {
background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 3px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);
}

</string>
      </property>
     </widget>
     <widget class="QComboBox" name="selectEditColor">
      <property name="geometry">
       <rect>
        <x>20</x>
        <y>306</y>
        <width>171</width>
        <height>41</height>
       </rect>
      </property>
      <property name="styleSheet">
       <string notr="true">background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 3px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);</string>
      </property>
      <item>
       <property name="text">
        <string>Синий</string>
       </property>
      </item>
      <item>
       <property name="text">
        <string>Красный</string>
       </property>
      </item>
      <item>
       <property name="text">
        <string>Жёлтый</string>
       </property>
      </item>
      <item>
       <property name="text">
        <string>Чёрный</string>
       </property>
      </item>
      <item>
       <property name="text">
        <string>Фиолетовый</string>
       </property>
      </item>
      <item>
       <property name="text">
        <string>Зелёный</string>
       </property>
      </item>
      <item>
       <property name="text">
        <string>Оранжевый</string>
       </property>
      </item>
      <item>
       <property name="text">
        <string>Розовый</string>
       </property>
      </item>
     </widget>
     <widget class="QPushButton" name="editGoalButton">
      <property name="geometry">
       <rect>
        <x>35</x>
        <y>355</y>
        <width>121</width>
        <height>51</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <family>Arial</family>
        <pointsize>10</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="cursor">
       <cursorShape>PointingHandCursor</cursorShape>
      </property>
      <property name="styleSheet">
       <string notr="true">QPushButton {
background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 10px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);
}

QPushButton:hover {
background-color: rgba(255,33,100,100);
}</string>
      </property>
      <property name="text">
       <string>Изменить</string>
      </property>
     </widget>
     <widget class="QPushButton" name="closeEditGoalWindowButton">
      <property name="geometry">
       <rect>
        <x>270</x>
        <y>10</y>
        <width>40</width>
        <height>40</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <pointsize>10</pointsize>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">QPushButton {
background-color: rgba(255,33,100,150);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 10px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);
}

QPushButton:hover {
background-color: rgba(255,33,100,100);
}</string>
      </property>
      <property name="text">
       <string>X</string>
      </property>
     </widget>
     <widget class="QLabel" name="editGoalResult">
      <property name="geometry">
       <rect>
        <x>0</x>
        <y>415</y>
        <width>321</width>
        <height>41</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <family>Arial</family>
        <pointsize>12</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">color:white;</string>
      </property>
      <property name="text">
       <string/>
      </property>
     </widget>
     <widget class="QLabel" name="goalEditLabel">
      <property name="geometry">
       <rect>
        <x>0</x>
        <y>70</y>
        <width>321</width>
        <height>41</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <family>Arial Black</family>
        <pointsize>14</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">color: white</string>
      </property>
      <property name="text">
       <string>Изменяемая цель</string>
      </property>
     </widget>
     <widget class="QSpinBox" name="selectEditGoal">
      <property name="geometry">
       <rect>
        <x>200</x>
        <y>70</y>
        <width>71</width>
        <height>41</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <pointsize>10</pointsize>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 3px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);</string>
      </property>
      <property name="minimum">
       <number>1</number>
      </property>
      <property name="maximum">
       <number>8</number>
      </property>
     </widget>
     <widget class="QPushButton" name="deleteGoalButton">
      <property name="geometry">
       <rect>
        <x>170</x>
        <y>355</y>
        <width>121</width>
        <height>51</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <family>Arial</family>
        <pointsize>10</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="cursor">
       <cursorShape>PointingHandCursor</cursorShape>
      </property>
      <property name="styleSheet">
       <string notr="true">QPushButton {
background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 10px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);
}

QPushButton:hover {
background-color: rgba(255,33,100,100);
}</string>
      </property>
      <property name="text">
       <string>Удалить</string>
      </property>
     </widget>
    </widget>
    <widget class="QWidget" name="progressGoalMenu" native="true">
     <property name="geometry">
      <rect>
       <x>329</x>
       <y>49</y>
       <width>321</width>
       <height>461</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">border-radius: 10px;
background-color: qlineargradient(spread:reflect, x1:0, y1:0.011, x2:1, y2:0, stop:0 rgba(189, 36, 126, 243), stop:1 rgba(255, 255, 255, 255));</string>
     </property>
     <widget class="QPushButton" name="closeProgressGoalWindowButton">
      <property name="geometry">
       <rect>
        <x>270</x>
        <y>10</y>
        <width>40</width>
        <height>40</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <pointsize>10</pointsize>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">QPushButton {
background-color: rgba(255,33,100,150);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 10px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);
}

QPushButton:hover {
background-color: rgba(255,33,100,100);
}</string>
      </property>
      <property name="text">
       <string>X</string>
      </property>
     </widget>
     <widget class="QTableWidget" name="goalsinformation">
      <property name="geometry">
       <rect>
        <x>0</x>
        <y>60</y>
        <width>321</width>
        <height>121</height>
       </rect>
      </property>
      <property name="styleSheet">
       <string notr="true">color: black;</string>
      </property>
     </widget>
     <widget class="QLabel" name="progressGoalResult">
      <property name="geometry">
       <rect>
        <x>0</x>
        <y>400</y>
        <width>321</width>
        <height>41</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <family>Arial</family>
        <pointsize>12</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">color:white;</string>
      </property>
      <property name="text">
       <string/>
      </property>
     </widget>
     <widget class="QPushButton" name="progressGoalButton">
      <property name="geometry">
       <rect>
        <x>60</x>
        <y>330</y>
        <width>211</width>
        <height>51</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <pointsize>10</pointsize>
       </font>
      </property>
      <property name="cursor">
       <cursorShape>PointingHandCursor</cursorShape>
      </property>
      <property name="styleSheet">
       <string notr="true">QPushButton {
background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 10px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);
}

QPushButton:hover {
background-color: rgba(255,33,100,100);
}</string>
      </property>
      <property name="text">
       <string>Добавить</string>
      </property>
     </widget>
     <widget class="QLabel" name="goalEditProgressLabel">
      <property name="geometry">
       <rect>
        <x>0</x>
        <y>190</y>
        <width>321</width>
        <height>41</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <family>Arial Black</family>
        <pointsize>14</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">color: white</string>
      </property>
      <property name="text">
       <string>Изменяемая цель</string>
      </property>
     </widget>
     <widget class="QLabel" name="goalSummProgressLabel">
      <property name="geometry">
       <rect>
        <x>0</x>
        <y>230</y>
        <width>321</width>
        <height>41</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <family>Arial Black</family>
        <pointsize>14</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">color: white;</string>
      </property>
      <property name="text">
       <string>Сумма</string>
      </property>
      <property name="alignment">
       <set>Qt::AlignLeading|Qt::AlignLeft|Qt::AlignVCenter</set>
      </property>
     </widget>
     <widget class="QLineEdit" name="entersSummGoalProgress">
      <property name="geometry">
       <rect>
        <x>12</x>
        <y>270</y>
        <width>301</width>
        <height>41</height>
       </rect>
      </property>
      <property name="styleSheet">
       <string notr="true">QLineEdit {
background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 3px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);
}

</string>
      </property>
     </widget>
     <widget class="QSpinBox" name="selectEditGoalProgress">
      <property name="geometry">
       <rect>
        <x>220</x>
        <y>190</y>
        <width>71</width>
        <height>41</height>
       </rect>
      </property>
      <property name="font">
       <font>
        <pointsize>10</pointsize>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 3px;
padding: 10px;
font-size: 10pt;
color: rgba(255, 255, 255, 200);</string>
      </property>
      <property name="minimum">
       <number>1</number>
      </property>
      <property name="maximum">
       <number>8</number>
      </property>
     </widget>
    </widget>
    <zorder>goalPlannerPlace</zorder>
    <zorder>goalPlannerLabel</zorder>
    <zorder>goalProgress1</zorder>
    <zorder>goalProgressLabel1</zorder>
    <zorder>goalProgressLabel2</zorder>
    <zorder>goalProgress2</zorder>
    <zorder>goalProgressLabel3</zorder>
    <zorder>goalProgress3</zorder>
    <zorder>goalProgressLabel4</zorder>
    <zorder>goalProgress4</zorder>
    <zorder>goalProgressLabel5</zorder>
    <zorder>goalProgress5</zorder>
    <zorder>goalProgressLabel6</zorder>
    <zorder>goalProgress6</zorder>
    <zorder>goalProgressLabel7</zorder>
    <zorder>goalProgress7</zorder>
    <zorder>goalProgressLabel8</zorder>
    <zorder>goalProgress8</zorder>
    <zorder>updateGoals</zorder>
    <zorder>createGoal</zorder>
    <zorder>editGoal</zorder>
    <zorder>goalNumber_1</zorder>
    <zorder>goalNumber_2</zorder>
    <zorder>goalNumber_3</zorder>
    <zorder>goalNumber_4</zorder>
    <zorder>goalNumber_5</zorder>
    <zorder>goalNumber_6</zorder>
    <zorder>goalNumber_7</zorder>
    <zorder>goalNumber_8</zorder>
    <zorder>addProgress</zorder>
    <zorder>addGoal</zorder>
    <zorder>editGoalMenu</zorder>
    <zorder>progressGoalMenu</zorder>
   </widget>
   <widget class="QWidget" name="currencyConvertatorMenu" native="true">
    <property name="geometry">
     <rect>
      <x>30</x>
      <y>160</y>
      <width>961</width>
      <height>550</height>
     </rect>
    </property>
    <widget class="QLabel" name="currencyConvertatorPlace">
     <property name="geometry">
      <rect>
       <x>0</x>
       <y>0</y>
       <width>961</width>
       <height>551</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">background-color: rgba(0, 0, 0, 150);
border-radius: 10px;</string>
     </property>
     <property name="text">
      <string/>
     </property>
    </widget>
    <widget class="QLabel" name="picture">
     <property name="geometry">
      <rect>
       <x>330</x>
       <y>75</y>
       <width>321</width>
       <height>201</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">border-image: url(:/Images/istockphoto-1436091608-170667a (1).jpg)</string>
     </property>
     <property name="text">
      <string/>
     </property>
    </widget>
    <widget class="QLabel" name="currencyConvertatorMenuLabel_2">
     <property name="geometry">
      <rect>
       <x>310</x>
       <y>20</y>
       <width>361</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>24</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Конвертатор валют</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLabel" name="convertibleCurrencyLabel">
     <property name="geometry">
      <rect>
       <x>60</x>
       <y>80</y>
       <width>251</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>12</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Конвертируемая валюта</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QComboBox" name="convertibleCurrency">
     <property name="geometry">
      <rect>
       <x>60</x>
       <y>120</y>
       <width>251</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <pointsize>8</pointsize>
       <weight>50</weight>
       <bold>false</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">background-color: rgb(58, 58, 60);
border: 2px solid rgba(255, 255, 255, 250);
color: white;
border-radius: 5px</string>
     </property>
     <item>
      <property name="text">
       <string>Австралийский доллар (AUD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Азербайджанский манат (AZN)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Албанский лек (ALL)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Алжирский динар (DZD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Ангольская кванза (AOA)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Аргентинское песо (ARS)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Армянский драм (AMD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Арубанский флорин (AWG)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Афганский афгани (AFN)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Багамский доллар (BSD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Бангладешская така (BDT)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Барбадосский доллар (BBD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Бахрейнский динар (BHD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Белизский доллар (BZD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Белорусский рубль (BYN)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Болгарский лев (BGN)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Боливийский боливиано (BOB)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Ботсванская пула (BWP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Бразильский реал (BRL)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Брунейский доллар (BND)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Бурундийский франк (BIF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Венгерский форинт (HUF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Восточнокарибский доллар (XCD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Вьетнамский донг (VND)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Гаитянский гурд (HTG)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Гайанский доллар (GYD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Гамбийский даласи (GMD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Ганский седи (GHS)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Гватемальский кетсаль (GTQ)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Гвинейский франк (GNF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Гондурасская лемпира (HNL)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Гонконгский доллар (HKD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Грузинский лари (GEL)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Датская крона (DKK)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Дирхам ОАЭ (AED)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Доллар Намибии (NAD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Доллар США (USD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Доллар Соломоновых Островов (SBD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Доллар Тринидада и Тобаго (TTD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Доллар Фиджи (FJD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Доминиканское песо (DOP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Евро (EUR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Египетский фунт (EGP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Жэньминьби (CNY)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Западноафриканский франк КФА (XOF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Индийская рупия (INR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Индонезийская рупия (IDR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Иорданский динар (JOD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Иранский риал (IRR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Исландская крона (ISK)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Йеменский риал (YER)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Казахстанский тенге (KZT)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Камбоджийский риель (KHR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Канадский доллар (CAD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Катарский риал (QAR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Кенийский шиллинг (KES)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Кина (PGK)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Киргизский сом (KGS)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Колумбийское песо (COP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Конвертируемая марка Боснии и Герцеговины (BAM)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Конголезский франк (CDF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Коста-риканский колон (CRC)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Кувейтский динар (KWD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Лаосский кип (LAK)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Либерийский доллар (LRD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Ливанский фунт (LBP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Ливийский динар (LYD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Лоти Лесото (LSL)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Маврикийская рупия (MUR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Мавританская угия (MRU)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Македонский денар (MKD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Малавийская квача (MWK)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Малагасийский ариари (MGA)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Малайзийский ринггит (MYR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Мальдивская руфия (MVR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Марокканский дирхам (MAD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Мексиканское песо (MXN)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Мозамбикский метикал (MZN)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Молдавский лей (MDL)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Мьянманский кьят (MMK)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Непальская рупия (NPR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Нигерийская найра (NGN)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Нидерландский антильский гульден (ANG)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Никарагуанская кордоба (NIO)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Новозеландский доллар (NZD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Новый израильский шекель (ILS)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Новый тайваньский доллар (TWD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Норвежская крона (NOK)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Оманский риал (OMR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Пакистанская рупия (PKR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Панамский бальбоа (PAB)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Парагвайский гуарани (PYG)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Патака Макао (MOP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Перуанский соль (PEN)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Российский рубль (RUB)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Румынский лей (RON)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Сальвадорский колон (SVC)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Саудовский риял (SAR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Свазилендский лилангени (SZL)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Сейшельская рупия (SCR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Суверенный боливар (VES)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Суданский фунт (SDG)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Суринамский доллар (SRD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Сьерра-леонский леоне (SLL)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Таджикский сомони (TJS)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Тайский бат (THB)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Танзанийский шиллинг (TZS)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Тонганская паанга (TOP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Тунисский динар (TND)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Турецкая лира (TRY)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Туркменский манат (TMT)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Угандийский шиллинг (UGX)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Узбекский сум (UZS)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Украинская гривна (UAH)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Уругвайское песо (UYU)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Филиппинское песо  (PHP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Франк Джибути (DJF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Франк КФП (XPF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Франк Комор (KMF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Франк Руанды (RWF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Фунт стерлингов (GBP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Центральноафриканский франк КФА (XAF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Чешская крона (CZK)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Чилийское песо (CLP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Шведская крона (SEK)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Швейцарский франк (CHF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Шри-ланкийская рупия (LKR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Эскудо Кабо-Верде (CVE)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Эфиопский быр (ETB)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Южноафриканский рэнд (ZAR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Южнокорейская вона (KRW)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Ямайский доллар (JMD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Японская иена (JPY)</string>
      </property>
     </item>
    </widget>
    <widget class="QComboBox" name="convertedCurrency">
     <property name="geometry">
      <rect>
       <x>670</x>
       <y>120</y>
       <width>251</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <pointsize>8</pointsize>
       <weight>50</weight>
       <bold>false</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">background-color: rgb(58, 58, 60);
border: 2px solid rgba(255, 255, 255, 250);
color: white;
border-radius: 5px</string>
     </property>
     <item>
      <property name="text">
       <string>Австралийский доллар (AUD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Азербайджанский манат (AZN)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Албанский лек (ALL)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Алжирский динар (DZD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Ангольская кванза (AOA)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Аргентинское песо (ARS)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Армянский драм (AMD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Арубанский флорин (AWG)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Афганский афгани (AFN)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Багамский доллар (BSD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Бангладешская така (BDT)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Барбадосский доллар (BBD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Бахрейнский динар (BHD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Белизский доллар (BZD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Белорусский рубль (BYN)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Болгарский лев (BGN)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Боливийский боливиано (BOB)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Ботсванская пула (BWP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Бразильский реал (BRL)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Брунейский доллар (BND)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Бурундийский франк (BIF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Венгерский форинт (HUF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Восточнокарибский доллар (XCD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Вьетнамский донг (VND)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Гаитянский гурд (HTG)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Гайанский доллар (GYD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Гамбийский даласи (GMD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Ганский седи (GHS)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Гватемальский кетсаль (GTQ)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Гвинейский франк (GNF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Гондурасская лемпира (HNL)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Гонконгский доллар (HKD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Грузинский лари (GEL)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Датская крона (DKK)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Дирхам ОАЭ (AED)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Доллар Намибии (NAD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Доллар США (USD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Доллар Соломоновых Островов (SBD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Доллар Тринидада и Тобаго (TTD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Доллар Фиджи (FJD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Доминиканское песо (DOP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Евро (EUR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Египетский фунт (EGP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Жэньминьби (CNY)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Западноафриканский франк КФА (XOF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Индийская рупия (INR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Индонезийская рупия (IDR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Иорданский динар (JOD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Иранский риал (IRR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Исландская крона (ISK)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Йеменский риал (YER)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Казахстанский тенге (KZT)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Камбоджийский риель (KHR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Канадский доллар (CAD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Катарский риал (QAR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Кенийский шиллинг (KES)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Кина (PGK)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Киргизский сом (KGS)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Колумбийское песо (COP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Конвертируемая марка Боснии и Герцеговины (BAM)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Конголезский франк (CDF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Коста-риканский колон (CRC)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Кувейтский динар (KWD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Лаосский кип (LAK)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Либерийский доллар (LRD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Ливанский фунт (LBP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Ливийский динар (LYD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Лоти Лесото (LSL)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Маврикийская рупия (MUR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Мавританская угия (MRU)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Македонский денар (MKD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Малавийская квача (MWK)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Малагасийский ариари (MGA)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Малайзийский ринггит (MYR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Мальдивская руфия (MVR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Марокканский дирхам (MAD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Мексиканское песо (MXN)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Мозамбикский метикал (MZN)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Молдавский лей (MDL)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Мьянманский кьят (MMK)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Непальская рупия (NPR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Нигерийская найра (NGN)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Нидерландский антильский гульден (ANG)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Никарагуанская кордоба (NIO)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Новозеландский доллар (NZD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Новый израильский шекель (ILS)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Новый тайваньский доллар (TWD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Норвежская крона (NOK)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Оманский риал (OMR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Пакистанская рупия (PKR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Панамский бальбоа (PAB)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Парагвайский гуарани (PYG)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Патака Макао (MOP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Перуанский соль (PEN)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Российский рубль (RUB)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Румынский лей (RON)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Сальвадорский колон (SVC)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Саудовский риял (SAR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Свазилендский лилангени (SZL)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Сейшельская рупия (SCR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Суверенный боливар (VES)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Суданский фунт (SDG)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Суринамский доллар (SRD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Сьерра-леонский леоне (SLL)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Таджикский сомони (TJS)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Тайский бат (THB)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Танзанийский шиллинг (TZS)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Тонганская паанга (TOP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Тунисский динар (TND)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Турецкая лира (TRY)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Туркменский манат (TMT)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Угандийский шиллинг (UGX)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Узбекский сум (UZS)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Украинская гривна (UAH)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Уругвайское песо (UYU)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Филиппинское песо  (PHP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Франк Джибути (DJF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Франк КФП (XPF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Франк Комор (KMF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Франк Руанды (RWF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Фунт стерлингов (GBP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Центральноафриканский франк КФА (XAF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Чешская крона (CZK)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Чилийское песо (CLP)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Шведская крона (SEK)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Швейцарский франк (CHF)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Шри-ланкийская рупия (LKR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Эскудо Кабо-Верде (CVE)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Эфиопский быр (ETB)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Южноафриканский рэнд (ZAR)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Южнокорейская вона (KRW)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Ямайский доллар (JMD)</string>
      </property>
     </item>
     <item>
      <property name="text">
       <string>Японская иена (JPY)</string>
      </property>
     </item>
    </widget>
    <widget class="QLabel" name="convertedCurrencyLabel">
     <property name="geometry">
      <rect>
       <x>670</x>
       <y>80</y>
       <width>251</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>12</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Получаемая валюта</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLabel" name="convertedAmountLabel">
     <property name="geometry">
      <rect>
       <x>670</x>
       <y>170</y>
       <width>251</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>12</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Полученная сумма</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLabel" name="convertibleAmountLabel">
     <property name="geometry">
      <rect>
       <x>60</x>
       <y>170</y>
       <width>251</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial Black</family>
       <pointsize>12</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;</string>
     </property>
     <property name="text">
      <string>Сумма</string>
     </property>
     <property name="alignment">
      <set>Qt::AlignCenter</set>
     </property>
    </widget>
    <widget class="QLineEdit" name="convertibleAmount">
     <property name="geometry">
      <rect>
       <x>60</x>
       <y>210</y>
       <width>251</width>
       <height>41</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">background-color: rgb(58, 58, 60);
border: 2px solid rgba(255, 255, 255, 250);
color: white;
border-radius: 5px</string>
     </property>
    </widget>
    <widget class="QLineEdit" name="convertedAmount">
     <property name="geometry">
      <rect>
       <x>670</x>
       <y>210</y>
       <width>251</width>
       <height>41</height>
      </rect>
     </property>
     <property name="styleSheet">
      <string notr="true">background-color: rgb(58, 58, 60);
border: 2px solid rgba(255, 255, 255, 250);
color: white;
border-radius: 5px</string>
     </property>
    </widget>
    <widget class="QPushButton" name="convertButton">
     <property name="geometry">
      <rect>
       <x>330</x>
       <y>330</y>
       <width>321</width>
       <height>61</height>
      </rect>
     </property>
     <property name="maximumSize">
      <size>
       <width>16777215</width>
       <height>16777209</height>
      </size>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>20</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">QPushButton {
background-color: rgb(58, 58, 60);
border: 2px solid rgba(255, 255, 255, 250);
color: white;
border-radius: 5px
}
QPushButton:hover {
background-color:rgba(58, 58, 60, 200);
}</string>
     </property>
     <property name="text">
      <string>Конвертировать</string>
     </property>
    </widget>
    <widget class="QLabel" name="convertLabelError">
     <property name="geometry">
      <rect>
       <x>330</x>
       <y>279</y>
       <width>321</width>
       <height>41</height>
      </rect>
     </property>
     <property name="font">
      <font>
       <family>Arial</family>
       <pointsize>11</pointsize>
       <weight>75</weight>
       <bold>true</bold>
      </font>
     </property>
     <property name="styleSheet">
      <string notr="true">color: white;
</string>
     </property>
     <property name="text">
      <string/>
     </property>
    </widget>
   </widget>
   <zorder>background</zorder>
   <zorder>mainMenuLabel</zorder>
   <zorder>currencyConverterLabel</zorder>
   <zorder>settingsLabel</zorder>
   <zorder>settings</zorder>
   <zorder>expenseManagerLabel</zorder>
   <zorder>revenueManagerLabel</zorder>
   <zorder>goalPlannerManager</zorder>
   <zorder>revenueManager</zorder>
   <zorder>expenseManager</zorder>
   <zorder>currencyConverter</zorder>
   <zorder>goalPlanner</zorder>
   <zorder>mainMenu</zorder>
   <zorder>MainMenu</zorder>
   <zorder>revenueManagerMenu</zorder>
   <zorder>expenseManagerMenu</zorder>
   <zorder>goalPlannerMenu</zorder>
   <zorder>currencyConvertatorMenu</zorder>
  </widget>
 </widget>
 <resources>
  <include location="res.qrc"/>
  <include location="../Users/miron/PycharmProjects/pythonProject/Resources.qrc"/>
 </resources>
 <connections/>
</ui>
"""
settingsTemplate = """<?xml version="1.0" encoding="UTF-8"?>
<ui version="4.0">
 <class>Form</class>
 <widget class="QWidget" name="Form">
  <property name="geometry">
   <rect>
    <x>0</x>
    <y>0</y>
    <width>551</width>
    <height>321</height>
   </rect>
  </property>
  <property name="minimumSize">
   <size>
    <width>551</width>
    <height>321</height>
   </size>
  </property>
  <property name="maximumSize">
   <size>
    <width>3232323</width>
    <height>16777215</height>
   </size>
  </property>
  <property name="windowTitle">
   <string>Form</string>
  </property>
  <widget class="QWidget" name="verticalLayoutWidget">
   <property name="geometry">
    <rect>
     <x>10</x>
     <y>10</y>
     <width>531</width>
     <height>301</height>
    </rect>
   </property>
   <layout class="QVBoxLayout" name="verticalLayout">
    <item>
     <widget class="QPushButton" name="closeSettings">
      <property name="font">
       <font>
        <family>Arial</family>
        <pointsize>12</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="toolTip">
       <string/>
      </property>
      <property name="toolTipDuration">
       <number>1</number>
      </property>
      <property name="statusTip">
       <string/>
      </property>
      <property name="styleSheet">
       <string notr="true">QPushButton {
background-color: rgba(0, 0, 0, 150);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 10px;

color: rgba(255, 255, 255, 200);
}

QPushButton:hover {
background-color: rgba(255,33,100,150);
}</string>
      </property>
      <property name="text">
       <string>Закрыть настройки</string>
      </property>
     </widget>
    </item>
    <item>
     <widget class="QLabel" name="settingText">
      <property name="font">
       <font>
        <family>Arial</family>
        <pointsize>16</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">background-color: rgba(0, 0, 0, 150);
border-radius: 10px;
color: white;</string>
      </property>
      <property name="text">
       <string>Настройки</string>
      </property>
     </widget>
    </item>
    <item>
     <layout class="QHBoxLayout" name="horizontalLayout">
      <item>
       <widget class="QLabel" name="disableAutomaticLoginLabel">
        <property name="font">
         <font>
          <family>Arial</family>
          <pointsize>10</pointsize>
          <weight>75</weight>
          <bold>true</bold>
         </font>
        </property>
        <property name="styleSheet">
         <string notr="true">background-color: rgba(0, 0, 0, 150);
border-radius: 10px;
color: white;</string>
        </property>
        <property name="text">
         <string>Отключить автоматический вход</string>
        </property>
       </widget>
      </item>
      <item>
       <widget class="QPushButton" name="disableAutomaticLoginButton">
        <property name="font">
         <font>
          <family>Arial</family>
          <pointsize>12</pointsize>
          <weight>75</weight>
          <bold>true</bold>
         </font>
        </property>
        <property name="styleSheet">
         <string notr="true">QPushButton {
background-color: rgba(0, 0, 0, 150);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 10px;

color: rgba(255, 255, 255, 200);
}

QPushButton:hover {
background-color: rgba(255,33,100,150);
}</string>
        </property>
        <property name="text">
         <string>Отключить</string>
        </property>
       </widget>
      </item>
     </layout>
    </item>
    <item>
     <widget class="QLabel" name="exitProgram">
      <property name="font">
       <font>
        <family>Arial</family>
        <pointsize>10</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="styleSheet">
       <string notr="true">background-color: rgba(0, 0, 0, 150);
border-radius: 10px;
color: white;</string>
      </property>
      <property name="text">
       <string>Выйти из программы</string>
      </property>
     </widget>
    </item>
    <item>
     <widget class="QPushButton" name="exitButton">
      <property name="font">
       <font>
        <family>Arial</family>
        <pointsize>12</pointsize>
        <weight>75</weight>
        <bold>true</bold>
       </font>
      </property>
      <property name="toolTip">
       <string/>
      </property>
      <property name="toolTipDuration">
       <number>1</number>
      </property>
      <property name="statusTip">
       <string/>
      </property>
      <property name="styleSheet">
       <string notr="true">QPushButton {
background-color: rgba(0, 0, 0, 150);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 10px;

color: rgba(255, 255, 255, 200);
}

QPushButton:hover {
background-color: rgba(255,33,100,150);
}</string>
      </property>
      <property name="text">
       <string>Выход</string>
      </property>
     </widget>
    </item>
   </layout>
  </widget>
  <widget class="QLabel" name="background">
   <property name="geometry">
    <rect>
     <x>10</x>
     <y>8</y>
     <width>531</width>
     <height>301</height>
    </rect>
   </property>
   <property name="styleSheet">
    <string notr="true">border-image:url(:/Images/SettingsBackground.jpg);
border-radius: 20px;
</string>
   </property>
   <property name="text">
    <string/>
   </property>
  </widget>
  <widget class="QLabel" name="border">
   <property name="geometry">
    <rect>
     <x>0</x>
     <y>0</y>
     <width>551</width>
     <height>321</height>
    </rect>
   </property>
   <property name="styleSheet">
    <string notr="true">background-color:rgb(44, 109, 168);
border-radius: 20px;</string>
   </property>
   <property name="text">
    <string/>
   </property>
  </widget>
  <zorder>border</zorder>
  <zorder>background</zorder>
  <zorder>verticalLayoutWidget</zorder>
 </widget>
 <resources>
  <include location="res.qrc"/>
  <include location="../Users/miron/PycharmProjects/pythonProject/Resources.qrc"/>
 </resources>
 <connections/>
</ui>
"""
errorDisableAutomaticLoginTemplate = """<?xml version="1.0" encoding="UTF-8"?>
<ui version="4.0">
 <class>Form</class>
 <widget class="QWidget" name="Form">
  <property name="geometry">
   <rect>
    <x>0</x>
    <y>0</y>
    <width>1214</width>
    <height>803</height>
   </rect>
  </property>
  <property name="windowTitle">
   <string>Form</string>
  </property>
  <widget class="QLabel" name="border">
   <property name="geometry">
    <rect>
     <x>520</x>
     <y>330</y>
     <width>321</width>
     <height>180</height>
    </rect>
   </property>
   <property name="styleSheet">
    <string notr="true">background-color:rgb(44, 109, 168);</string>
   </property>
   <property name="text">
    <string/>
   </property>
  </widget>
  <widget class="QLabel" name="placeForText">
   <property name="geometry">
    <rect>
     <x>530</x>
     <y>340</y>
     <width>301</width>
     <height>161</height>
    </rect>
   </property>
   <property name="font">
    <font>
     <family>Arial</family>
     <pointsize>16</pointsize>
     <weight>75</weight>
     <bold>true</bold>
    </font>
   </property>
   <property name="styleSheet">
    <string notr="true">background-color: rgba(0, 0, 0, 150);

color: white;</string>
   </property>
   <property name="text">
    <string/>
   </property>
  </widget>
  <widget class="QPushButton" name="exitFromError">
   <property name="geometry">
    <rect>
     <x>640</x>
     <y>440</y>
     <width>71</width>
     <height>41</height>
    </rect>
   </property>
   <property name="font">
    <font>
     <family>Arial</family>
     <pointsize>12</pointsize>
     <weight>75</weight>
     <bold>true</bold>
    </font>
   </property>
   <property name="styleSheet">
    <string notr="true">QPushButton {
background-color: rgba(0,0,0,0);
border: 2px solid rgba(255, 33, 100, 230);
border-radius: 10px;

color: rgba(255, 255, 255, 200);
}

QPushButton:hover {
background-color: rgba(255,33,100,150);
}</string>
   </property>
   <property name="text">
    <string>Ок</string>
   </property>
  </widget>
  <widget class="QLabel" name="errorText">
   <property name="geometry">
    <rect>
     <x>560</x>
     <y>360</y>
     <width>261</width>
     <height>51</height>
    </rect>
   </property>
   <property name="font">
    <font>
     <family>Arial</family>
     <pointsize>10</pointsize>
     <weight>75</weight>
     <bold>true</bold>
    </font>
   </property>
   <property name="styleSheet">
    <string notr="true">color: white;</string>
   </property>
   <property name="text">
    <string>Автоматический вход уже выключен</string>
   </property>
  </widget>
  <widget class="QLabel" name="background">
   <property name="geometry">
    <rect>
     <x>530</x>
     <y>340</y>
     <width>301</width>
     <height>161</height>
    </rect>
   </property>
   <property name="styleSheet">
    <string notr="true">border-image:url(:/Images/ErrorBackground.jpg);</string>
   </property>
   <property name="text">
    <string/>
   </property>
  </widget>
  <zorder>border</zorder>
  <zorder>background</zorder>
  <zorder>placeForText</zorder>
  <zorder>exitFromError</zorder>
  <zorder>errorText</zorder>
 </widget>
 <resources>
  <include location="res.qrc"/>
  <include location="Resources.qrc"/>
 </resources>
 <connections/>
</ui>
"""


class RegistrationWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        f = io.StringIO(ReLogWindow)
        uic.loadUi(f, self)
        self.BackgroundUpdate()
        self.WindowTransparency()
        con = sqlite3.connect('Databases/RecordedLoginAndPassword')
        cur = con.cursor()
        data = cur.execute("""SELECT * FROM loginpassword WHERE id = '1'""").fetchall()
        if data:
            self.enterLogin.setText(data[-1][1])
            self.enterPassword.setText(data[-1][2])
            self.reminderButton.setChecked(True)
        con.close()
        self.logButton.clicked.connect(self.log)
        self.regButton.clicked.connect(self.reg)
        self.closeWindowButton.clicked.connect(self.closeWindow)

    def BackgroundUpdate(self):
        self.label.setStyleSheet("""border-image:url(:/Images/ReLogBackground.jpg);
         border-radius: 20px;""")

    def WindowTransparency(self):
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground, True)
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)

    def log(self):
        if self.enterPassword.text() == '' and self.enterLogin.text() == '':
            self.ReLogResult.setText('Введите логин и пароль')
        elif self.enterPassword.text() != '' and self.enterLogin.text() == '':
            self.ReLogResult.setText('Введите логин')
        elif self.enterPassword.text() == '' and self.enterLogin.text() != '':
            self.ReLogResult.setText('Введите пароль')
        else:
            self.login = self.enterLogin.text()
            con = sqlite3.connect('Databases/LoginsAndPasswords')
            cur = con.cursor()
            LoginAndPassword = cur.execute(f"""SELECT * FROM logpass WHERE
             login = '{self.enterLogin.text()}'""").fetchall()
            if LoginAndPassword != [] and LoginAndPassword[0][1] == self.enterPassword.text():
                con1 = sqlite3.connect('Databases/RecordedLoginAndPassword')
                cur1 = con1.cursor()
                if self.reminderButton.isChecked():
                    login = self.enterLogin.text()
                    passw = self.enterPassword.text()
                    cur1.execute(f"INSERT INTO loginpassword (id,login,password) VALUES('1', '{login}', '{passw}')")
                    con1.commit()
                    con1.close
                    self.GoEnd()
                else:
                    cur1.execute("""DELETE FROM loginpassword WHERE id = '1'""")
                    a = cur1.execute("""SELECT * FROM loginpassword""").fetchall()
                    con1.commit()
                    con1.close()
                    self.GoEnd()
            else:
                self.ReLogResult.setText('Неверный логин или пароль')
            con.close()

    def reg(self):
        if self.enterPassword.text() == '' and self.enterLogin.text() == '':
            self.ReLogResult.setText('Введите логин и пароль')
        elif self.enterPassword.text() != '' and self.enterLogin.text() == '':
            self.ReLogResult.setText('Введите логин')
        elif self.enterPassword.text() == '' and self.enterLogin.text() != '':
            self.ReLogResult.setText('Введите пароль')
        else:
            con = sqlite3.connect('Databases/LoginsAndPasswords')
            cur = con.cursor()
            logins = cur.execute(f"""SELECT login FROM logpass WHERE login = '{self.enterLogin.text()}'""").fetchall()
            passwords = (cur.execute(f"""SELECT password FROM logpass
             WHERE password = '{self.enterPassword.text()}'""").fetchall())
            if logins and not passwords:
                self.ReLogResult.setText('Данный логин занят')
                self.enterLogin.setText('')
            elif not logins and passwords:
                self.ReLogResult.setText('Данный пароль занят')
                self.enterPassword.setText('')
            elif logins and passwords:
                self.ReLogResult.setText('Данные логин и пароль заняты')
                self.enterLogin.setText('')
                self.enterPassword.setText('')
            elif len(self.enterLogin.text()) > 14:
                self.ReLogResult.setText('Логин слишком длинный')
                self.enterLogin.setText('')
            else:
                self.login = self.enterLogin.text()
                passw = self.enterPassword.text()
                date = datetime.now().date()
                con1 = sqlite3.connect('Databases/UsersInformat')
                cur1 = con1.cursor()
                cur1.execute(f"INSERT INTO inf (username, balance, numberofauthorizations, registrationdate,"
                             f" daysintheapp, numberoftransactions) VALUES('{self.login}', '0', '0', '{date}', '0', '0')")
                con1.commit()
                cur.execute(f"INSERT INTO logpass (login,password) VALUES('{self.login}', '{passw}')")
                con.commit()
                self.ReLogResult.setText('Вы успешно зарегистрировались!')
            con.close()

    def GoEnd(self):
        self.hide()
        self.app2 = LoadingWindow(self.login)
        self.app2.show()

    def closeWindow(self):
        sys.exit(app.exec_())


class LoadingWindow(QMainWindow):
    def __init__(self, login):
        self.login = login
        super().__init__()
        f = io.StringIO(LoadWindow)
        uic.loadUi(f, self)
        self.endLoad = False
        self.n = 100
        self.initUi()

    def initUi(self):
        f = io.StringIO(LoadWindow)
        uic.loadUi(f, self)
        self.BackgroundUpdate()
        self.WindowTransparency()
        self.timer = QtCore.QTimer()
        with open('OtherFiles/LoadCheckbox.txt', mode='r', encoding='UTF-8') as file:
            data = file.read()
        if data == 'True':
            self.checkBox.setChecked(True)
        self.timer.setInterval(100)
        self.timer.timeout.connect(self.run)
        QtCore.QTimer.singleShot(100, self.timer.start)
        self.pushButton.clicked.connect(self.GoEnd)

    def run(self):
        for i in range(self.n):
            time.sleep(0.01)
            self.progressBar.setValue(i + 1)
        self.timer.stop()
        if self.checkBox.isChecked():
            self.GoEnd()

    def BackgroundUpdate(self):
        self.label.setStyleSheet("""border-image:url(:/Images/LoadingBackground.jpg);
         border-radius: 20px;""")

    def WindowTransparency(self):
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground, True)
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)

    def GoEnd(self):
        if self.checkBox.isChecked():
            with open('OtherFiles/LoadCheckbox.txt', mode='w', encoding='UTF-8') as file:
                file.write('True')
        con = sqlite3.connect('Databases/UsersInformat')
        cur = con.cursor()
        count = cur.execute(f"SELECT numberofauthorizations FROM inf WHERE username = '{self.login}'").fetchall()[0][0]
        count = int(count) + 1
        cur.execute(f"UPDATE inf SET numberofauthorizations = {count} WHERE username = '{self.login}'")
        con.commit()
        con.close()
        self.hide()
        self.app2 = MainWindow(self.login)
        self.app2.show()


class MainWindow(QMainWindow):
    def __init__(self, login):
        self.loginText = login
        self.colors = {'Синий': 'blue', 'Красный': 'red', 'Жёлтый': 'yellow', 'Чёрный': 'black', 'Фиолетовый': 'violet',
                       'Зелёный': 'green', 'Оранжевый': 'orange', 'Розовый': 'pink', 'Голубой': 'yellow'}
        self.data = [['1', 'Не используется', '', '0', '0', 'blue'], ['2', 'Не используется', '', '0', '0', 'blue'],
                     ['3', 'Не используется', '', '0', '0', 'blue'], ['4', 'Не используется', '', '0', '0', 'blue'],
                     ['5', 'Не используется', '', '0', '0', 'blue'], ['6', 'Не используется', '', '0', '0', 'blue'],
                     ['7', 'Не используется', '', '0', '0', 'blue'], ['8', 'Не используется', '', '0', '0', 'blue']]
        super().__init__()
        f = io.StringIO(MainWindowTemplate)
        uic.loadUi(f, self)
        self.BackgroundUpdate()
        self.WindowTransparency()
        self.UpdateInformation()
        self.hideMenu()
        self.sortRevenueParameter.addItem('Источник')
        self.MainMenu.show()
        self.addRevenueErrorLabel.setText('')
        self.addExpenseErrorLabel.setText('')
        con = sqlite3.connect('Databases/UsersInformat')
        cur = con.cursor()
        data = cur.execute(f"SELECT numberofauthorizations FROM inf WHERE username "
                           f"= '{self.loginText}'").fetchall()[0][0]
        con.close()
        if data == 1:
            self.curr_image = QImage('Avatars/Иконка.png').scaled(122, 122)
            self.pixmapcopy = self.curr_image.copy()
            self.pixmap = QPixmap.fromImage(self.curr_image)
            self.pixmap.save(f"Avatars/{self.loginText}.png", "png");
            self.avatar.setPixmap(self.pixmap)
        else:
            self.curr_image = QImage(f"Avatars/{self.loginText}.png")
            self.pixmapcopy = self.curr_image.copy()
            self.pixmap = QPixmap.fromImage(self.curr_image)
            self.pixmap.save(f"Avatars/{self.loginText}.png", "png");
            self.avatar.setPixmap(self.pixmap)
        self.mainMenu.clicked.connect(self.OpenMainMenu)
        self.revenueManager.clicked.connect(self.OpenRevenueManager)
        self.expenseManager.clicked.connect(self.OpenExpenseManager)
        self.goalPlanner.clicked.connect(self.OpenGoalPlannerMenu)
        self.settings.clicked.connect(self.OpenSettings)
        self.createGoal.clicked.connect(self.OpenAddGoal)
        self.currencyConverter.clicked.connect(self.OpenCurrencyConverter)
        self.addRevenue.clicked.connect(self.createRevenue)
        self.addExpense.clicked.connect(self.createExpense)
        self.updateRevenueButton.clicked.connect(self.UpdateRevenueTransactions)
        self.updateExpenseButton.clicked.connect(self.UpdateExpenseTransactions)
        self.closeAddGoalWindowButton.clicked.connect(self.closeAddGoalWindow)
        self.closeEditGoalWindowButton.clicked.connect(self.closeEditGoalWindow)
        self.addGoalButton.clicked.connect(self.addGoalFunction)
        self.editGoalButton.clicked.connect(self.editGoalFunction)
        self.editGoal.clicked.connect(self.OpenEditGoal)
        self.updateGoals.clicked.connect(self.UpdateGoalsFunction)
        self.deleteGoalButton.clicked.connect(self.DeleteGoalFunction)
        self.closeProgressGoalWindowButton.clicked.connect(self.closeProgressGoalWindow)
        self.addProgress.clicked.connect(self.OpenGoalProgress)
        self.progressGoalButton.clicked.connect(self.addProgressFunction)
        self.convertButton.clicked.connect(self.convert)
        self.editAvatar.clicked.connect(self.EditAvatarFunction)

    def EditAvatarFunction(self):
        self.name = QFileDialog.getOpenFileName(self, 'Выбрать картинку', '')[0]
        self.curr_image = QImage(self.name).scaled(122, 122)
        self.pixmapcopy = self.curr_image.copy()
        self.pixmap = QPixmap.fromImage(self.curr_image)
        self.pixmap.save(f"Avatars/{self.loginText}.png", "png");
        self.avatar.setPixmap(self.pixmap)

    def hideMenu(self):
        self.expenseManagerMenu.hide()
        self.revenueManagerMenu.hide()
        self.goalPlannerMenu.hide()
        self.currencyConvertatorMenu.hide()
        self.MainMenu.hide()

    def BackgroundUpdate(self):
        self.background.setStyleSheet("""border-image:url(:/Images/BackgroudMainWindow.jpg);
         border-radius: 20px;""")

    def WindowTransparency(self):
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground, True)
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)

    def UpdateInformation(self):
        self.login.setText(self.loginText)
        con1 = sqlite3.connect('Databases/UsersInformat')
        cur1 = con1.cursor()
        data = cur1.execute(f"SELECT balance, registrationdate, numberoftransactions FROM inf WHERE username "
                            f"= '{self.loginText}'").fetchall()
        self.balance.setText(f'Ваш баланс: {data[0][0]}')
        date = datetime.now().date()
        date1 = data[0][1]
        date1 = '/'.join(date1.split('-')[::-1])
        date2 = datetime.strptime(f'{date1}', '%d/%m/%Y')
        self.userInformation.setText(f'Информация:\n\nДата регистрации:\n\n{data[0][1]}\n\nВ приложении:'
                                     f'\n\n{(date - date2.date()).days} дней\n\nКоличество транзакций:\n\n{data[0][2]}')
        con1.close()
        self.recentTransactions.clear()
        self.recentTransactions.setRowCount(0)
        wb = load_workbook('ДенежныеТранзакции.xlsx')
        sheetnames = wb.sheetnames
        self.recentTransactions.setColumnCount(5)
        self.recentTransactions.setHorizontalHeaderLabels(['Доходы\nРасходы', 'Сумма', 'Источник', 'Категория', 'Дата'])
        if self.loginText in sheetnames:
            ws = wb[self.loginText]
            indexation = 1
            data = []
            while ws[f'E{indexation}'].value is not None:
                data.append([ws[f'E{indexation}'].value, ws[f'A{indexation}'].value, ws[f'B{indexation}'].value,
                             ws[f'C{indexation}'].value, ws[f'D{indexation}'].value])
                indexation += 1
            for i, row in enumerate(data[::-1][0:10]):
                self.recentTransactions.setRowCount(self.recentTransactions.rowCount() + 1)
                self.recentTransactions.setItem(i, 0, QTableWidgetItem(str(row[0])))
                self.recentTransactions.setItem(i, 2, QTableWidgetItem(str(row[2])))
                self.recentTransactions.setItem(i, 1, QTableWidgetItem(str(row[1])))
                self.recentTransactions.setItem(i, 3, QTableWidgetItem(str(row[3])))
                self.recentTransactions.setItem(i, 4, QTableWidgetItem(str(row[4])))
            wb.close()


    def OpenMainMenu(self):
        self.hideMenu()
        self.UpdateInformation()
        self.MainMenu.show()

    def OpenRevenueManager(self):
        self.hideMenu()
        self.UpdateRevenueTransactions()
        self.revenueManagerMenu.show()

    def OpenAddGoal(self):
        self.hideMenu()
        self.OpenGoalPlannerMenu()
        self.addGoal.show()

    def OpenEditGoal(self):
        self.hideMenu()
        self.OpenGoalPlannerMenu()
        self.editGoalMenu.show()

    def OpenCurrencyConverter(self):
        self.hideMenu()
        self.currencyConvertatorMenu.show()

    def OpenGoalPlannerMenu(self):
        try:
            with open(f'Goals/goals{self.loginText}.txt', mode='r', encoding='utf-8') as file:
                goals = file.readlines()
        except:
            with open(f'Goals/goals{self.loginText}.txt', mode='w', encoding='UTF-8') as file:
                for goal in self.data:
                    file.writelines(';'.join(goal) + '\n')
        self.hideMenu()
        self.UpdateGoalsFunction()
        self.addGoal.hide()
        self.editGoalMenu.hide()
        self.progressGoalMenu.hide()
        self.goalPlannerMenu.show()

    def OpenGoalProgress(self):
        self.goalsinformation.setColumnCount(3)
        self.goalsinformation.setHorizontalHeaderLabels(['Номер', 'Цель', 'Прогресс'])
        with open(f'Goals/goals{self.loginText}.txt', mode='r', encoding='utf-8') as file:
            goals = file.readlines()
            goals = [goal[:len(goal) - 1].split(';') if '\n' in goal else goal.split(';') for goal in goals]
        for i, row in enumerate(goals):
            self.goalsinformation.setRowCount(self.goalsinformation.rowCount() + 1)
            self.goalsinformation.setItem(i, 0, QTableWidgetItem(str(row[0])))
            self.goalsinformation.setItem(i, 1, QTableWidgetItem(str(row[3])))
            self.goalsinformation.setItem(i, 2, QTableWidgetItem(str(row[4])))
        self.progressGoalMenu.show()

    def createRevenue(self):
        try:
            a = int(self.revenueSummaEnter.text())
            if a < 0:
                self.addRevenueErrorLabel.setText('Ошибка: Укажите сумму без минуса')
            else:
                if (self.revenueSummaEnter.text() == '' or self.revenueSourceEnter.text() == ''
                        or self.revenueCategoryEnter.text() == ''):
                    self.addRevenueErrorLabel.setText('Ошибка: Заполните все поля')
                else:
                    wb = load_workbook('ДенежныеТранзакции.xlsx')
                    sheetnames = wb.sheetnames
                    if self.loginText in sheetnames:
                        ws = wb[self.loginText]
                    else:
                        ws = wb.create_sheet(self.loginText)
                    ws.append([self.revenueSummaEnter.text(), self.revenueSourceEnter.text(),
                               self.revenueCategoryEnter.text(),
                               datetime.now(), 'Д'])
                    wb.save('ДенежныеТранзакции.xlsx')
                    wb.close()
                    con = sqlite3.connect('Databases/UsersInformat')
                    cur = con.cursor()
                    balance = cur.execute(f"SELECT balance FROM inf WHERE username = '{self.loginText}'").fetchall()[0][
                        0]
                    count = cur.execute(f"SELECT numberoftransactions FROM inf WHERE username"
                                        f" = '{self.loginText}'").fetchall()[0][0]
                    count = int(count) + 1
                    balance = int(balance) + int(self.revenueSummaEnter.text())
                    cur.execute(f"UPDATE inf SET numberoftransactions = {count} WHERE username = '{self.loginText}'")
                    cur.execute(f"UPDATE inf SET balance = {balance} WHERE username = '{self.loginText}'")
                    con.commit()
                    con.close()
                    self.revenueSummaEnter.setText('')
                    self.revenueSourceEnter.setText('')
                    self.revenueCategoryEnter.setText('')
                    self.addRevenueErrorLabel.setText('')
        except Exception:
            self.addRevenueErrorLabel.setText('Ошибка: Сумма должна \nсостоять из цифр')

    def UpdateRevenueTransactions(self):
        self.revenueTransactions.clear()
        self.revenueTransactions.setRowCount(0)
        wb = load_workbook('ДенежныеТранзакции.xlsx')
        sheetnames = wb.sheetnames
        if self.loginText in sheetnames:
            ws = wb[self.loginText]
            indexation = 1
            data = []
            while ws[f'E{indexation}'].value is not None:
                if ws[f'E{indexation}'].value == 'Д':
                    data.append([ws[f'A{indexation}'].value, ws[f'B{indexation}'].value, ws[f'C{indexation}'].value,
                                 ws[f'D{indexation}'].value])
                indexation += 1
            self.revenueTransactions.setColumnCount(4)
            self.revenueTransactions.setHorizontalHeaderLabels(['Сумма', 'Источник', 'Категория', 'Дата'])
            if self.sortRevenueParameter.currentText() == 'Дата':
                data = data[::-1]
            elif self.sortRevenueParameter.currentText() == 'Категория':
                data = sorted(data, key=lambda x: (x[2], int(x[0])))[::-1]
            elif self.sortRevenueParameter.currentText() == 'От максимального':
                data = sorted(data, key=lambda x: int(x[0]))[::-1]
            elif self.sortRevenueParameter.currentText() == 'От минимального':
                data = sorted(data, key=lambda x: int(x[0]))
            elif self.sortRevenueParameter.currentText() == 'Источник':
                data = sorted(data, key=lambda x: (x[1], int(x[0])))[::-1]
            for i, row in enumerate(data):
                self.revenueTransactions.setRowCount(self.revenueTransactions.rowCount() + 1)
                self.revenueTransactions.setItem(i, 0, QTableWidgetItem(str(row[0])))
                self.revenueTransactions.setItem(i, 2, QTableWidgetItem(str(row[2])))
                self.revenueTransactions.setItem(i, 1, QTableWidgetItem(str(row[1])))
                self.revenueTransactions.setItem(i, 3, QTableWidgetItem(str(row[3])))
            wb.close()

    def UpdateExpenseTransactions(self):
        self.expenseTransactions.clear()
        self.expenseTransactions.setRowCount(0)
        wb = load_workbook('ДенежныеТранзакции.xlsx')
        sheetnames = wb.sheetnames
        if self.loginText in sheetnames:
            ws = wb[self.loginText]
            indexation = 1
            data = []
            while ws[f'E{indexation}'].value is not None:
                if ws[f'E{indexation}'].value == 'Р':
                    data.append([ws[f'A{indexation}'].value, ws[f'B{indexation}'].value, ws[f'C{indexation}'].value,
                                 ws[f'D{indexation}'].value])
                indexation += 1
            self.expenseTransactions.setColumnCount(4)
            self.expenseTransactions.setHorizontalHeaderLabels(['Сумма', 'Источник', 'Категория', 'Дата'])
            if self.sortExpenseParameter.currentText() == 'Дата':
                data = data[::-1]
            elif self.sortExpenseParameter.currentText() == 'Категория':
                data = sorted(data, key=lambda x: (x[2], int(x[0])))[::-1]
            elif self.sortExpenseParameter.currentText() == 'От максимальной суммы':
                data = sorted(data, key=lambda x: int(x[0]))[::-1]
            elif self.sortExpenseParameter.currentText() == 'От минимальной суммы':
                data = sorted(data, key=lambda x: int(x[0]))
            elif self.sortExpenseParameter.currentText() == 'Источник':
                data = sorted(data, key=lambda x: (x[1], int(x[0])))[::-1]
            for i, row in enumerate(data):
                self.expenseTransactions.setRowCount(self.expenseTransactions.rowCount() + 1)
                self.expenseTransactions.setItem(i, 0, QTableWidgetItem(str(row[0])))
                self.expenseTransactions.setItem(i, 2, QTableWidgetItem(str(row[2])))
                self.expenseTransactions.setItem(i, 1, QTableWidgetItem(str(row[1])))
                self.expenseTransactions.setItem(i, 3, QTableWidgetItem(str(row[3])))

            wb.close()

    def createExpense(self):
        try:
            a = int(self.expenseSummaEnter.text())
            if a < 0:
                self.addExpenseErrorLabel.setText('Ошибка: Укажите сумму без минуса')
            else:
                if (self.expenseSummaEnter.text() == '' or self.expenseSourceEnter.text() == ''
                        or self.expenseCategoryEnter.text() == ''):
                    self.addExpenseErrorLabel.setText('Ошибка: Заполните все поля')
                else:
                    wb = load_workbook('ДенежныеТранзакции.xlsx')
                    sheetnames = wb.sheetnames
                    if self.loginText in sheetnames:
                        ws = wb[self.loginText]
                    else:
                        ws = wb.create_sheet(self.loginText)
                    ws.append(
                        [self.expenseSummaEnter.text(), self.expenseSourceEnter.text(),
                         self.expenseCategoryEnter.text(),
                         datetime.now(), 'Р'])
                    wb.save('ДенежныеТранзакции.xlsx')
                    wb.close()
                    con = sqlite3.connect('Databases/UsersInformat')
                    cur = con.cursor()
                    balance = cur.execute(f"SELECT balance FROM inf WHERE username = '{self.loginText}'").fetchall()[0][
                        0]
                    count = cur.execute(f"SELECT numberoftransactions FROM inf WHERE username"
                                        f" = '{self.loginText}'").fetchall()[0][0]
                    count = int(count) + 1
                    balance = int(balance) - int(self.expenseSummaEnter.text())
                    cur.execute(f"UPDATE inf SET numberoftransactions = {count} WHERE username = '{self.loginText}'")
                    cur.execute(f"UPDATE inf SET balance = {balance} WHERE username = '{self.loginText}'")
                    con.commit()
                    con.close()
                    self.expenseSummaEnter.setText('')
                    self.expenseSourceEnter.setText('')
                    self.expenseCategoryEnter.setText('')
                    self.addExpenseErrorLabel.setText('')
        except Exception:
            self.addExpenseErrorLabel.setText('Ошибка: Сумма должна \nсостоять из цифр')

    def addGoalFunction(self):
        try:
            if self.enterGoalTarget.text() == '' or self.enterGoalName.text() == '':
                self.addGoalResult.setText('Ошибка: Заполните все поля')
            else:
                a = int(self.enterGoalTarget.text())
                if a < 0:
                    self.addGoalResult.setText('Ошибка: Укажите цель без минуса')
                else:
                    with open(f'Goals/goals{self.loginText}.txt', mode='r', encoding='utf-8') as file:
                        goals = file.readlines()
                        goals = [goal[:len(goal) - 1].split(';') if '\n' in goal else goal.split(';') for goal in goals]
                        count = 0
                        flag = 0
                        changed_goals = []
                        for goal in goals:
                            if goal[1] == 'Не используется':
                                if flag == 0:
                                    goal[1] = 'Используется'
                                    goal[2] = self.enterGoalName.text()
                                    goal[3] = self.enterGoalTarget.text()
                                    goal[5] = self.selectColor.currentText()
                                    flag = 1
                            else:
                                count += 1
                            changed_goals.append(goal)
                        if count == 8:
                            self.addExpenseErrorLabel.setText('Ошибка: Все цели заняты')
                        else:
                            with open(f'Goals/goals{self.loginText}.txt', mode='w', encoding='UTF-8') as file:
                                for goal in changed_goals:
                                    file.writelines(';'.join(goal) + '\n')
                        self.addGoalResult.setText('')
                        self.enterGoalTarget.setText('')
                        self.enterGoalName.setText('')
                        self.selectColor.setCurrentIndex(0)
                        self.addGoal.hide()
        except Exception:
            self.addGoalResult.setText('Ошибка: Цель должна \nсостоять из цифр')

    def addProgressFunction(self):
        try:
            if self.entersSummGoalProgress.text() != '':
                a = int(self.entersSummGoalProgress.text())
                if a < 0:
                    self.progressGoalResult.setText('Ошибка: Укажите цель без минуса')
                    return
            with open(f'Goals/goals{self.loginText}.txt', mode='r', encoding='utf-8') as file:
                goals = file.readlines()
                goals = [goal[:len(goal) - 1].split(';') if '\n' in goal else goal.split(';') for goal in goals]
            if goals[self.selectEditGoalProgress.value() - 1][1] == 'Не используется':
                self.progressGoalResult.setText('Ошибка: Данная цель \nне используется')
                return
            else:
                if int(goals[self.selectEditGoalProgress.value() - 1][4]) + \
                    int(self.entersSummGoalProgress.text()) > int(goals[self.selectEditGoalProgress.value() - 1][3]):
                    self.progressGoalResult.setText('Ошибка: Слишком большая \nсумма для цели')
                    return
                goals[self.selectEditGoalProgress.value() - 1][4] \
                 = str(int(goals[self.selectEditGoalProgress.value() - 1][4]) + int(self.entersSummGoalProgress.text()))
                print(goals)
                with open(f'Goals/goals{self.loginText}.txt', mode='w', encoding='UTF-8') as file:
                    for goal in goals:
                        file.writelines(';'.join(goal) + '\n')
            self.progressGoalResult.setText('')
            self.entersSummGoalProgress.setText('')
            self.selectEditGoalProgress.setValue(1)
            self.progressGoalMenu.hide()
        except Exception:
            self.progressGoalResult.setText('Ошибка: Цель должна \nсостоять из цифр')

    def editGoalFunction(self):
        try:
            if self.enterEditGoalTarget.text() != '':
                a = int(self.enterEditGoalTarget.text())
                if a < 0:
                    self.editGoalResult.setText('Ошибка: Укажите цель без минуса')
                    return
            with open(f'Goals/goals{self.loginText}.txt', mode='r', encoding='utf-8') as file:
                goals = file.readlines()
                goals = [goal[:len(goal) - 1].split(';') if '\n' in goal else goal.split(';') for goal in goals]
            if goals[self.selectEditGoal.value() - 1][1] == 'Не используется':
                self.editGoalResult.setText('Ошибка: Данная цель не используется')
                return
            else:
                if self.enterEditGoalName.text() != '' and self.enterEditGoalTarget.text() != '':
                    goals[self.selectEditGoal.value() - 1][2] = self.enterEditGoalName.text()
                    goals[self.selectEditGoal.value() - 1][3] = self.enterEditGoalTarget.text()
                    goals[self.selectEditGoal.value() - 1][5] = self.selectEditColor.currentText()
                elif self.enterEditGoalName.text() != '':
                    goals[self.selectEditGoal.value() - 1][2] = self.enterEditGoalName.text()
                    goals[self.selectEditGoal.value() - 1][5] = self.selectEditColor.currentText()
                elif self.enterEditGoalTarget.text() != '':
                    goals[self.selectEditGoal.value() - 1][3] = self.enterEditGoalTarget.text()
                    goals[self.selectEditGoal.value() - 1][5] = self.selectEditColor.currentText()
                else:
                    goals[self.selectEditGoal.value() - 1][5] = self.selectEditColor.currentText()
            with open(f'Goals/goals{self.loginText}.txt', mode='w', encoding='UTF-8') as file:
                for goal in goals:
                    file.writelines(';'.join(goal) + '\n')
            self.editGoalResult.setText('')
            self.enterEditGoalTarget.setText('')
            self.enterEditGoalName.setText('')
            self.selectEditGoal.setValue(1)
            self.selectEditColor.setCurrentIndex(0)
            self.editGoalMenu.hide()
        except Exception:
            self.editGoalResult.setText('Ошибка: Цель должна \nсостоять из цифр')

    def UpdateGoalsFunction(self):
        with open(f'Goals/goals{self.loginText}.txt', mode='r', encoding='utf-8') as file:
            goals = file.readlines()
            goals = [goal[:len(goal) - 1].split(';') if '\n' in goal else goal.split(';') for goal in goals]
        for goal in goals:
            if goal[1] == 'Используется':
                color = self.colors[goal[5]]
                styleSheet = Template("""QProgressBar{
                                    background-color: rgb(124, 113, 116);
                                    border-radius: 12px;
                                    color: white;
                                    text-align: center;
                                    border: 2px solid rgba(255, 255, 255, 250);
                                    }
                                    QProgressBar::chunk{
                                    border-radius: 12px;
                                    background-color: $color
                                    }""")
                if goal[0] == '1':
                    self.goalProgress1.setValue(math.floor(round((int(goal[4])) / int(goal[3]), 2) * 100))
                    self.goalProgressLabel1.setText(goal[2])
                    self.goalProgress1.setStyleSheet(styleSheet.substitute(color=color))
                if goal[0] == '2':
                    self.goalProgress2.setValue(math.floor(round((int(goal[4])) / int(goal[3]), 2) * 100))
                    self.goalProgressLabel2.setText(goal[2])
                    self.goalProgress2.setStyleSheet(styleSheet.substitute(color=color))
                if goal[0] == '3':
                    self.goalProgress3.setValue(math.floor(round((int(goal[4])) / int(goal[3]), 2) * 100))
                    self.goalProgressLabel3.setText(goal[2])
                    self.goalProgress3.setStyleSheet(styleSheet.substitute(color=color))
                if goal[0] == '4':
                    self.goalProgress4.setValue(math.floor(round((int(goal[4])) / int(goal[3]), 2) * 100))
                    self.goalProgressLabel4.setText(goal[2])
                    self.goalProgress4.setStyleSheet(styleSheet.substitute(color=color))
                if goal[0] == '5':
                    self.goalProgress5.setValue(math.floor(round((int(goal[4])) / int(goal[3]), 2) * 100))
                    self.goalProgressLabel5.setText(goal[2])
                    self.goalProgress5.setStyleSheet(styleSheet.substitute(color=color))
                if goal[0] == '6':
                    self.goalProgress6.setValue(math.floor(round((int(goal[4])) / int(goal[3]), 2) * 100))
                    self.goalProgressLabel6.setText(goal[2])
                    self.goalProgress6.setStyleSheet(styleSheet.substitute(color=color))
                if goal[0] == '7':
                    self.goalProgress7.setValue(math.floor(round((int(goal[4])) / int(goal[3]), 2) * 100))
                    self.goalProgressLabel7.setText(goal[2])
                    self.goalProgress7.setStyleSheet(styleSheet.substitute(color=color))
                if goal[0] == '8':
                    self.goalProgress8.setValue(math.floor(round((int(goal[4])) / int(goal[3]), 2) * 100))
                    self.goalProgressLabel8.setText(goal[2])
                    self.goalProgress8.setStyleSheet(styleSheet.substitute(color=color))
            else:
                if goal[0] == '1':
                    self.goalProgress1.setValue(0)
                    self.goalProgressLabel1.setText('Не используется')
                if goal[0] == '2':
                    self.goalProgress2.setValue(0)
                    self.goalProgressLabel2.setText('Не используется')
                if goal[0] == '3':
                    self.goalProgress3.setValue(0)
                    self.goalProgressLabel3.setText('Не используется')
                if goal[0] == '4':
                    self.goalProgress4.setValue(0)
                    self.goalProgressLabel4.setText('Не используется')
                if goal[0] == '5':
                    self.goalProgress5.setValue(0)
                    self.goalProgressLabel5.setText('Не используется')
                if goal[0] == '6':
                    self.goalProgress6.setValue(0)
                    self.goalProgressLabel6.setText('Не используется')
                if goal[0] == '7':
                    self.goalProgress7.setValue(0)
                    self.goalProgressLabel7.setText('Не используется')
                if goal[0] == '8':
                    self.goalProgress8.setValue(0)
                    self.goalProgressLabel8.setText('Не используется')

    def DeleteGoalFunction(self):
        with open(f'Goals/goals{self.loginText}.txt', mode='r', encoding='utf-8') as file:
            goals = file.readlines()
            goals = [goal[:len(goal) - 1].split(';') if '\n' in goal else goal.split(';') for goal in goals]
        if goals[self.selectEditGoal.value() - 1][1] == 'Не используется':
            self.editGoalResult.setText('Ошибка: Данная цель не используется')
            return
        else:
            goals[self.selectEditGoal.value() - 1][1] = 'Не используется'
            goals[self.selectEditGoal.value() - 1][2] = ''
            goals[self.selectEditGoal.value() - 1][3] = '0'
            goals[self.selectEditGoal.value() - 1][4] = '0'
            goals[self.selectEditGoal.value() - 1][5] = 'blue'
        with open(f'Goals/goals{self.loginText}.txt', mode='w', encoding='UTF-8') as file:
            for goal in goals:
                file.writelines(';'.join(goal) + '\n')
        self.editGoalResult.setText('')
        self.enterEditGoalTarget.setText('')
        self.enterEditGoalName.setText('')
        self.selectEditGoal.setValue(1)
        self.selectEditColor.setCurrentIndex(0)
        self.editGoalMenu.hide()

    def convert(self):
        first_currency = self.convertibleCurrency.currentText()
        first_currency = first_currency.replace('(', '')
        first_currency = first_currency.replace(')', '')
        first_currency = first_currency.split()[-1]
        second_currency = self.convertedCurrency.currentText()
        second_currency = second_currency.replace('(', '')
        second_currency = second_currency.replace(')', '')
        second_currency = second_currency.split()[-1]
        amount = self.convertibleAmount.text()
        if self.convertibleAmount.text() == '':
            self.convertLabelError.setText('Ошибка: Введите сумму конвертации')
            return
        try:
            amount = int(self.convertibleAmount.text())
            if amount < 0:
                self.convertLabelError.setText('Ошибка: Укажите сумму без минуса')
                return
            if first_currency == second_currency:
                self.convertedAmount.setText(str(amount))
                self.convertLabelError.setText('')
                return
            try:
                result = currency.convert(first_currency, second_currency, amount)
                self.convertedAmount.setText(str(result))
            except Exception:
                fromFirstCurrencyToRubles = currency.convert(first_currency, 'RUB', 1)
                fromSecondCurrencyToRubles = currency.convert(second_currency, 'RUB', 1)
                self.convertedAmount.setText(str(round(float(fromFirstCurrencyToRubles) / float(fromSecondCurrencyToRubles)
                                             * amount, 2)))
        except Exception:
            self.convertLabelError.setText('Ошибка: Сумма должна состоять из цифр')
            return
        self.convertLabelError.setText('')

    def OpenExpenseManager(self):
        self.hideMenu()
        self.UpdateExpenseTransactions()
        self.expenseManagerMenu.show()

    def OpenSettings(self):
        self.app4 = SettingsWindow()
        self.app4.show()

    def closeAddGoalWindow(self):
        self.addGoal.hide()

    def closeEditGoalWindow(self):
        self.editGoalMenu.hide()

    def closeProgressGoalWindow(self):
        self.progressGoalMenu.hide()


class SettingsWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        f = io.StringIO(settingsTemplate)
        uic.loadUi(f, self)
        self.setWindowTitle('Settings')
        self.BackgroundUpdate()
        self.disableAutomaticLoginButton.clicked.connect(self.DisableAutomaticLoginFunction)
        self.exitButton.clicked.connect(self.ExitProgram)
        self.closeSettings.clicked.connect(self.closeSettingsFunction)

    def resizeEvent(self, event):
        new_size = event.size()
        self.border.resize(new_size.width(), new_size.height())
        self.background.resize(new_size.width() - 20, new_size.height() - 20)
        super().resizeEvent(event)
    def BackgroundUpdate(self):
        self.background.setStyleSheet("""border-image:url(:/Images/SettingsBackground.jpg);
         border-radius: 20px;""")

    def WindowTransparency(self):
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground, True)
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)

    def DisableAutomaticLoginFunction(self):
        with open('OtherFiles/LoadCheckbox.txt', mode='r', encoding='UTF-8') as file:
            condition = file.readline()
        if condition == 'True':
            with open('OtherFiles/LoadCheckbox.txt', mode='w', encoding='UTF-8') as file:
                file.write('False')
        else:
            self.app5 = ErrorDisableAutomaticLoginWindow()
            self.app5.show()

    def ExitProgram(self):
        sys.exit(app.exec_())

    def closeSettingsFunction(self):
        self.hide()


class ErrorDisableAutomaticLoginWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        f = io.StringIO(errorDisableAutomaticLoginTemplate)
        uic.loadUi(f, self)
        self.BackgroundUpdate()
        self.WindowTransparency()
        self.exitFromError.clicked.connect(self.closeError)

    def BackgroundUpdate(self):
        self.background.setStyleSheet("""border-image:url(:/Images/ErrorBackground.jpg);""")

    def WindowTransparency(self):
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground, True)
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)

    def closeError(self):
        self.hide()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = RegistrationWindow()
    ex.show()
    sys.exit(app.exec_())
